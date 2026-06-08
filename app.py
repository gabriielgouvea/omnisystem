import hashlib
import io
import json
import os
import re
import secrets
import shutil
import time as _time
import zipfile
from datetime import date, datetime, timedelta
from functools import wraps
import fitz  # PyMuPDF
from flask import Flask, request, jsonify, render_template, send_from_directory, send_file, session, redirect, url_for, make_response
from pypdf import PdfReader, PdfWriter
from pathlib import Path

app = Flask(__name__)
app.permanent_session_lifetime = timedelta(days=30)

BASE_DIR    = Path(__file__).parent
UPLOAD_DIR  = BASE_DIR / "uploads"
OUTPUT_DIR  = BASE_DIR / "outputs"
ARCHIVE_DIR = BASE_DIR / "archive"
MAPPINGS_FILE = BASE_DIR / "mappings.json"
BRANDS_FILE       = BASE_DIR / "brands.json"
HISTORY_FILE      = BASE_DIR / "history.json"
TRACKING_LOG_FILE   = BASE_DIR / "tracking_log.json"
TRACKING_INDEX_FILE = BASE_DIR / "tracking_index.json"
USERS_FILE        = BASE_DIR / "users.json"
AUDITORIA_FILE    = BASE_DIR / "auditoria.json"
SECRET_KEY_FILE   = BASE_DIR / ".secret_key"

if SECRET_KEY_FILE.exists():
    app.secret_key = SECRET_KEY_FILE.read_text().strip()
else:
    _key = secrets.token_hex(32)
    SECRET_KEY_FILE.write_text(_key)
    app.secret_key = _key

app.config["MAX_CONTENT_LENGTH"] = 1024 * 1024 * 1024  # 1 GB

@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "Arquivo muito grande. Limite: 200MB"}), 413

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
ARCHIVE_DIR.mkdir(exist_ok=True)

BRAND_DISPLAY_DEFAULT = {
    "pura":  "Creatina Black Skull Pura",
    "turbo": "Creatina Black Skull Turbo",
    "dux":   "Creatina DUX",
    "roupa": "Roupas",
}
STANDARD_ORDER = ["pura", "turbo", "dux", "roupa"]

BR_STATES = {"AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG",
             "PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"}

# Full state names â†’ abbreviation (longer names listed first to avoid partial matches)
BR_STATE_NAMES = [
    ("MATO GROSSO DO SUL", "MS"), ("MATO GROSSO", "MT"),
    ("ESPIRITO SANTO", "ES"),     ("ESPÃRITO SANTO", "ES"),
    ("RIO GRANDE DO NORTE", "RN"),("RIO GRANDE DO SUL", "RS"),
    ("RIO DE JANEIRO", "RJ"),     ("SANTA CATARINA", "SC"),
    ("SAO PAULO", "SP"),          ("SÃƒO PAULO", "SP"),
    ("MINAS GERAIS", "MG"),       ("DISTRITO FEDERAL", "DF"),
    ("ACRE", "AC"),               ("ALAGOAS", "AL"),
    ("AMAPA", "AP"),              ("AMAPÃ", "AP"),
    ("AMAZONAS", "AM"),           ("BAHIA", "BA"),
    ("CEARA", "CE"),              ("CEARÃ", "CE"),
    ("GOIAS", "GO"),              ("GOIÃS", "GO"),
    ("MARANHAO", "MA"),           ("MARANHÃƒO", "MA"),
    ("PARA", "PA"),               ("PARÃ", "PA"),
    ("PARAIBA", "PB"),            ("PARAÃBA", "PB"),
    ("PARANA", "PR"),             ("PARANÃ", "PR"),
    ("PERNAMBUCO", "PE"),
    ("PIAUI", "PI"),              ("PIAUÃ", "PI"),
    ("RONDONIA", "RO"),           ("RONDÃ”NIA", "RO"),
    ("RORAIMA", "RR"),            ("SERGIPE", "SE"),
    ("TOCANTINS", "TO"),
]

BR_HOLIDAYS_FIXED = {
    (1,  1): "ConfraternizaÃ§Ã£o Universal",
    (4, 21): "Tiradentes",
    (5,  1): "Dia do Trabalho",
    (9,  7): "IndependÃªncia do Brasil",
    (10,12): "Nossa Senhora Aparecida",
    (11, 2): "Finados",
    (11,15): "ProclamaÃ§Ã£o da RepÃºblica",
    (12,25): "Natal",
}

# Easter-based variable holidays pre-computed for 2025-2028
# Easter: 2025=Apr 20, 2026=Apr 5, 2027=Mar 28, 2028=Apr 16
BR_HOLIDAYS_VARIABLE = {
    2025: {(3,3):"Segunda de Carnaval",(3,4):"TerÃ§a de Carnaval",(3,5):"Quarta de Cinzas",
           (4,18):"PaixÃ£o de Cristo",(4,20):"PÃ¡scoa",(6,19):"Corpus Christi"},
    2026: {(2,16):"Segunda de Carnaval",(2,17):"TerÃ§a de Carnaval",(2,18):"Quarta de Cinzas",
           (4,3):"PaixÃ£o de Cristo",(4,5):"PÃ¡scoa",(6,4):"Corpus Christi"},
    2027: {(3,1):"Segunda de Carnaval",(3,2):"TerÃ§a de Carnaval",(3,3):"Quarta de Cinzas",
           (3,26):"PaixÃ£o de Cristo",(3,28):"PÃ¡scoa",(5,27):"Corpus Christi"},
    2028: {(2,28):"Segunda de Carnaval",(2,29):"TerÃ§a de Carnaval",(3,1):"Quarta de Cinzas",
           (4,14):"PaixÃ£o de Cristo",(4,16):"PÃ¡scoa",(6,15):"Corpus Christi"},
}

# â”€â”€ Mappings â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def load_mappings():
    if MAPPINGS_FILE.exists():
        with open(MAPPINGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_mappings(mappings):
    with open(MAPPINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(mappings, f, ensure_ascii=False, indent=2)


def load_brands():
    if BRANDS_FILE.exists():
        with open(BRANDS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_brands(brands):
    with open(BRANDS_FILE, "w", encoding="utf-8") as f:
        json.dump(brands, f, ensure_ascii=False, indent=2)


def get_brand_display():
    result = dict(BRAND_DISPLAY_DEFAULT)
    result.update(load_brands())
    return result


def load_history():
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_history(history):
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def load_tracking_log():
    if TRACKING_LOG_FILE.exists():
        with open(TRACKING_LOG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_tracking_log(log):
    with open(TRACKING_LOG_FILE, "w", encoding="utf-8") as f:
        json.dump(log, f, ensure_ascii=False, indent=2)


def load_tracking_index():
    if TRACKING_INDEX_FILE.exists():
        with open(TRACKING_INDEX_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_tracking_index(index):
    with open(TRACKING_INDEX_FILE, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)



def mapping_key(produto, sku):
    return f"{sku.strip()}|{_normalize_titulo(produto)}"


_TRACKING_RE = re.compile(r'\bBR[A-Z0-9]{8,25}\b')


def extract_tracking_numbers(page):
    """Return unique BR... tracking numbers found on a label page."""
    return list(set(_TRACKING_RE.findall(page.get_text("text"))))


def _normalize_titulo(t):
    """Normalise product title for loose comparison."""
    t = t.lower().strip()
    t = re.sub(r'[^\w\s]', '', t)
    return re.sub(r'\s+', ' ', t)


def detect_sku_title_conflicts(filenames, mappings, retirada_filenames=None):
    """Return SKUs that appear in the PDFs under a title not present in mappings,
    but whose base SKU IS mapped under a different title."""
    retirada_set = set(retirada_filenames or [])
    mapped_skus = {}  # base_sku -> list of (mk, info)
    for mk, info in mappings.items():
        if "|" in mk:
            base = mk.split("|", 1)[0]
            mapped_skus.setdefault(base, []).append((mk, info))

    conflicts = {}
    for filename in filenames:
        pdf_path = UPLOAD_DIR / filename
        doc = fitz.open(str(pdf_path))
        _cl_fn = extract_checklist_combined if filename in retirada_set else extract_checklist
        for page_num in range(len(doc)):
            for item in _cl_fn(doc[page_num]):
                sku   = item["sku"].strip()
                mk    = mapping_key(item["produto"], item["sku"])
                titulo = re.sub(r"\s+", " ", item["produto"]).strip()
                if mk in mappings:
                    continue  # known product, no conflict
                if sku not in mapped_skus or sku in conflicts:
                    continue
                # Same base SKU is mapped but under a different title
                existing = mapped_skus[sku]
                conflicts[sku] = {
                    "sku": sku,
                    "titulo_encontrado": titulo,
                    "titulo_salvo": existing[0][1].get("titulo", existing[0][0].split("|", 1)[-1]),
                    "categoria_atual": existing[0][1].get("categoria", "?"),
                    "kit_size": existing[0][1].get("kit_size", 1),
                }
        doc.close()
    return list(conflicts.values())


# â”€â”€ PDF extraction â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# Checklist column x-boundaries (consistent across all Shopee label PDFs)
COL_NUM_MAX     = 50
COL_PRODUTO_MAX = 120
COL_SKU_MAX     = 170
COL_VAR_MAX     = 215

def word_column(x):
    if x < COL_NUM_MAX:       return "num"
    elif x < COL_PRODUTO_MAX: return "produto"
    elif x < COL_SKU_MAX:     return "sku"
    elif x < COL_VAR_MAX:     return "variacao"
    else:                     return "quantidade"

def extract_checklist(page):
    """Extract checklist items from one label page using word positions."""
    words = page.get_text("words")
    checklist_y = None
    for w in words:
        x0, y0, x1, y1, word, *_ = w
        if "Checklist" in word:
            checklist_y = y0
            break
    if checklist_y is None:
        return []
    header_y = None
    for w in words:
        x0, y0, x1, y1, word, *_ = w
        if y0 > checklist_y and word.strip() == "Quantidade":
            header_y = y0
            break
    if header_y is None:
        return []
    data = []
    for w in words:
        x0, y0, x1, y1, word, *_ = w
        if y0 > header_y + 2:
            data.append({"x": x0, "y": y0, "word": word, "col": word_column(x0)})
    if not data:
        return []
    y_groups = {}
    for d in data:
        y_key = round(d["y"] / 3) * 3
        y_groups.setdefault(y_key, []).append(d)
    items = []
    current = None
    for y_key in sorted(y_groups):
        row = y_groups[y_key]
        num_words  = [d["word"] for d in row if d["col"] == "num"      and d["word"].isdigit()]
        prod_words = [d["word"] for d in row if d["col"] == "produto"]
        sku_words  = [d["word"] for d in row if d["col"] == "sku"]
        var_words  = [d["word"] for d in row if d["col"] == "variacao"]
        qty_words  = [d["word"] for d in row if d["col"] == "quantidade"]
        if num_words:
            if current is not None:
                items.append(current)
            qty_raw = qty_words[0] if qty_words else "0"
            qty = int(qty_raw) if qty_raw.isdigit() else 0
            current = {
                "produto":    " ".join(prod_words),
                "sku":        sku_words[0] if sku_words else "",
                "variacao":   " ".join(var_words),
                "quantidade": qty,
            }
        elif current is not None:
            if prod_words:
                current["produto"] += " " + " ".join(prod_words)
            if var_words and not current["variacao"]:
                current["variacao"] = " ".join(var_words)
    if current is not None:
        items.append(current)
    for item in items:
        item["produto"] = re.sub(r"\s+", " ", item["produto"]).strip()
    return items


def extract_checklist_combined(page):
    """Extrai checklist de pÃ¡ginas onde etiqueta e checklist estÃ£o lado a lado
    (formato RETIRADA PELO COMPRADOR da Shopee).
    Detecta as posiÃ§Ãµes reais das colunas pelo cabeÃ§alho para nÃ£o depender de
    offsets fixos (as colunas ficam comprimidas na metade direita da pÃ¡gina)."""
    words = page.get_text("words")
    # Encontra "Checklist" para ancorar a busca
    checklist_y = None
    for w in words:
        x0, y0, x1, y1, word, *_ = w
        if "Checklist" in word:
            checklist_y = y0
            break
    if checklist_y is None:
        return []
    # Encontra "Quantidade" (cabeÃ§alho da Ãºltima coluna) abaixo do tÃ­tulo
    header_y = None
    qty_header_x = None
    for w in words:
        x0, y0, x1, y1, word, *_ = w
        if y0 > checklist_y and word.strip() == "Quantidade":
            header_y = y0
            qty_header_x = x0
            break
    if header_y is None:
        return []
    # Mapeia posiÃ§Ã£o X real de cada coluna a partir do cabeÃ§alho
    col_xs = {}
    for w in words:
        x0, y0, x1, y1, word, *_ = w
        if abs(y0 - header_y) > 4:
            continue
        wc = word.strip()
        if wc == "#":
            col_xs["num"] = x0
        elif wc == "Produto":
            col_xs["produto"] = x0
        elif wc == "SKU":
            col_xs["sku"] = x0
        elif "Varia" in wc:
            col_xs["variacao"] = x0
        elif wc == "Quantidade":
            col_xs["quantidade"] = x0
    if "num" not in col_xs or "quantidade" not in col_xs:
        return []
    x_min = col_xs["num"] - 5
    # Classificador dinÃ¢mico: atribui cada x Ã  coluna mais prÃ³xima Ã  esquerda
    col_order = sorted((v, k) for k, v in col_xs.items())
    def col_of(x):
        result = col_order[0][1]
        for col_x, col_name in col_order:
            if x >= col_x - 5:
                result = col_name
        return result
    data = []
    for w in words:
        x0, y0, x1, y1, word, *_ = w
        if y0 > header_y + 2 and x0 >= x_min:
            data.append({"x": x0, "y": y0, "word": word, "col": col_of(x0)})
    if not data:
        return []
    y_groups = {}
    for d in data:
        y_key = round(d["y"] / 3) * 3
        y_groups.setdefault(y_key, []).append(d)
    items = []
    current = None
    for y_key in sorted(y_groups):
        row = y_groups[y_key]
        num_words  = [d["word"] for d in row if d["col"] == "num"      and d["word"].isdigit()]
        prod_words = [d["word"] for d in row if d["col"] == "produto"]
        sku_words  = [d["word"] for d in row if d["col"] == "sku"]
        var_words  = [d["word"] for d in row if d["col"] == "variacao"]
        qty_words  = [d["word"] for d in row if d["col"] == "quantidade"]
        if num_words:
            if current is not None:
                items.append(current)
            qty_raw = qty_words[0] if qty_words else "0"
            current = {
                "produto":    " ".join(prod_words),
                "sku":        sku_words[0] if sku_words else "",
                "variacao":   " ".join(var_words),
                "quantidade": int(qty_raw) if qty_raw.isdigit() else 0,
            }
        elif current is not None:
            if prod_words:
                current["produto"] += " " + " ".join(prod_words)
            if var_words and not current["variacao"]:
                current["variacao"] = " ".join(var_words)
    if current is not None:
        items.append(current)
    for item in items:
        item["produto"] = re.sub(r"\s+", " ", item["produto"]).strip()
    return items


def get_unknown_items_combined(pdf_path, mappings):
    doc = fitz.open(str(pdf_path))
    seen = {}
    for page_num in range(len(doc)):
        for item in extract_checklist_combined(doc[page_num]):
            key = mapping_key(item["produto"], item["sku"])
            if key not in mappings and key not in seen:
                seen[key] = {
                    "produto": item["produto"], "sku": item["sku"], "key": key,
                    "file": pdf_path.name, "page": page_num + 1,
                }
    doc.close()
    return list(seen.values())


def extract_declaration_items(page):
    """Extrai os itens da tabela 'IDENTIFICAÇÃO DOS BENS' de uma página de
    Declaração de Conteúdo (formato Shopee/Correios). Colunas:
    Nº | CÓDIGO (SKU) | DESCRIÇÃO DO PRODUTO | VARIAÇÃO | QTD | VALOR.
    Retorna [{produto, sku, quantidade}]."""
    words = page.get_text("words")
    ident_y = next((w[1] for w in words if "IDENTIFICA" in w[4].upper()), None)
    if ident_y is None:
        return []
    # Detecta posições X das colunas pelo cabeçalho (logo abaixo de "IDENTIFICAÇÃO")
    cols = {}
    header_y = None
    for w in words:
        if not (ident_y < w[1] < ident_y + 22):
            continue
        t = w[4].strip().upper()
        if "DIGO" in t:        cols["sku"] = w[0]; header_y = w[1]
        elif "DESCRI" in t:    cols["produto"] = w[0]
        elif "VARIA" in t:     cols["variacao"] = w[0]
        elif t == "QTD":       cols["qtd"] = w[0]
        elif "VALOR" in t:     cols["valor"] = w[0]
    if "produto" not in cols or "qtd" not in cols or header_y is None:
        return []
    sku_x   = cols.get("sku", 35)
    prod_x  = cols["produto"]
    qtd_x   = cols["qtd"]
    var_x   = cols.get("variacao", qtd_x - 90)
    valor_x = cols.get("valor", qtd_x + 25)

    def col_of(x):
        if x < sku_x - 3:   return "num"
        if x < prod_x - 3:  return "sku"
        if x < var_x - 3:   return "produto"
        if x < qtd_x - 3:   return "variacao"
        if x < valor_x - 3: return "qtd"
        return "valor"

    # Onde a tabela termina (Totais / Peso / DECLARAÇÃO)
    stop_y = None
    for w in words:
        if w[1] <= header_y + 3:
            continue
        tu = w[4].strip().upper()
        if tu in ("TOTAIS", "TOTAL") or tu.startswith("PESO") or "DECLARA" in tu:
            stop_y = w[1]; break

    rows = [w for w in words
            if w[1] > header_y + 3 and (stop_y is None or w[1] < stop_y - 2)]
    y_groups = {}
    for w in rows:
        y_groups.setdefault(round(w[1] / 3) * 3, []).append(w)

    items, current = [], None
    for yk in sorted(y_groups):
        by_col = {}
        for w in sorted(y_groups[yk], key=lambda w: w[0]):
            by_col.setdefault(col_of(w[0]), []).append(w[4])
        num = by_col.get("num", [])
        if num and num[0].strip().isdigit():
            if current:
                items.append(current)
            qv = by_col.get("qtd", [])
            q  = int(qv[0]) if qv and qv[0].strip().isdigit() else 1
            current = {
                "sku":        " ".join(by_col.get("sku", [])).strip(),
                "produto":    " ".join(by_col.get("produto", [])).strip(),
                "quantidade": q,
            }
        elif current and "produto" in by_col:
            current["produto"] += " " + " ".join(by_col["produto"])
    if current:
        items.append(current)
    for it in items:
        it["produto"] = re.sub(r"\s+", " ", it["produto"]).strip()
    return items


def get_unknown_items_declaracao(pdf_path, mappings):
    """Declaração: páginas ímpares (0,2,..)=etiqueta, pares (1,3,..)=declaração."""
    doc = fitz.open(str(pdf_path))
    seen = {}
    for i in range(1, len(doc), 2):     # páginas de declaração
        for item in extract_declaration_items(doc[i]):
            key = mapping_key(item["produto"], item["sku"])
            if key not in mappings and key not in seen:
                seen[key] = {
                    "produto": item["produto"], "sku": item["sku"], "key": key,
                    "file": pdf_path.name, "page": i + 1,
                }
    doc.close()
    return list(seen.values())


_CEP_RE = re.compile(r'^\d{5}-\d{3}$')

def extract_destination_state(page):
    """Extract destination state UF from a Shopee label page.

    Shopee labels write the full state name (e.g. 'SÃ£o Paulo', 'Bahia') in
    the address. The name may wrap across two lines, so we concatenate all
    words near the CEP into one text block and search for known state names.
    """
    words = page.get_text("words")
    page_h = page.rect.height
    dest_limit = page_h * 0.55  # destination section is the top ~55%

    # Word list restricted to destination section
    wlist = sorted(
        [(round(y0), w.strip(), round(x0))
         for x0, y0, x1, y1, w, *_ in words
         if w.strip() and y0 < dest_limit],
        key=lambda t: (t[0], t[2])
    )

    def _search_text(text):
        """Return UF if text contains a known state name or abbreviation."""
        tu = text.upper()
        for name, uf in BR_STATE_NAMES:
            if name in tu:
                return uf
        for token in tu.split():
            if len(token) == 2 and token in BR_STATES:
                return token
            for sep in ('/', '-'):
                if sep in token:
                    tail = token.rsplit(sep, 1)[-1]
                    if len(tail) == 2 and tail in BR_STATES:
                        return tail
        return None

    # Strategy 1: find destination CEP, combine ALL words in Â±50pt window
    cep_ys = [y for y, w, _ in wlist if _CEP_RE.match(w)]
    for cy in cep_ys:
        nearby = [w for y, w, _ in wlist if cy - 50 <= y <= cy + 15]
        uf = _search_text(" ".join(nearby))
        if uf:
            return uf

    # Fallback: combine all destination words and search
    uf = _search_text(" ".join(w for _, w, _ in wlist))
    return uf


def extract_store_name(pdf_path):
    """Extract store name from REMETENTE field on first label page."""
    doc = fitz.open(str(pdf_path))
    store = "Desconhecida"
    for page_num in range(min(3, len(doc))):
        words = doc[page_num].get_text("words")
        remetente_y = None
        for w in words:
            x0, y0, x1, y1, word, *_ = w
            if word.strip() == "REMETENTE" and x0 < 100:
                remetente_y = y0
                break
        if remetente_y is None:
            continue
        name_words = sorted(
            [(x0, word) for x0, y0, x1, y1, word, *_ in words
             if remetente_y + 2 < y0 < remetente_y + 14 and x0 < 200],
            key=lambda w: w[0]
        )
        if name_words:
            store = " ".join(w[1] for w in name_words).strip()
            break
    doc.close()
    return store


# â”€â”€ Classification â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def classify_page_items(items, mappings):
    unknown, category_units, category_set = [], {}, set()
    for item in items:
        key = mapping_key(item["produto"], item["sku"])
        if key not in mappings:
            unknown.append({"produto": item["produto"], "sku": item["sku"], "key": key})
        else:
            info = mappings[key]
            cat  = info["categoria"]
            kit  = info.get("kit_size", 1)
            category_set.add(cat)
            category_units[cat] = category_units.get(cat, 0) + kit * item["quantidade"]
    return {"unknown": unknown, "category_units": category_units, "category_set": list(category_set)}


def determine_output_pdf(items, mappings):
    for item in items:
        if mapping_key(item["produto"], item["sku"]) not in mappings:
            return None
    res      = classify_page_items(items, mappings)
    cat_set  = set(res["category_set"])
    cat_units = res["category_units"]
    if "roupa" in cat_set:
        return "roupas" if cat_set == {"roupa"} else "variacoes"
    if len(cat_set) > 1:
        return "variacoes"
    if len(cat_set) == 1:
        cat = list(cat_set)[0]
        return f"{cat}_{cat_units[cat]}"
    return "variacoes"


def check_label_format(pdf_path):
    """Sample up to 15 pages to detect if the Shopee label layout changed.
    Returns a list of issue strings, empty if everything looks normal.
    """
    doc = fitz.open(str(pdf_path))
    checklist_text_pages = 0
    items_found_pages    = 0
    sample = min(15, len(doc))
    for page_num in range(sample):
        page = doc[page_num]
        if "Checklist" in page.get_text("text"):
            checklist_text_pages += 1
            if extract_checklist(page):
                items_found_pages += 1
    doc.close()
    issues = []
    if checklist_text_pages > 0 and items_found_pages == 0:
        issues.append(
            f"{checklist_text_pages} pÃ¡gina(s) com 'Checklist' encontradas, "
            "mas nenhum item extraÃ­do â€” as colunas parecem estar em posiÃ§Ãµes diferentes."
        )
    elif checklist_text_pages > 0 and items_found_pages < checklist_text_pages * 0.4:
        issues.append(
            f"SÃ³ {items_found_pages} de {checklist_text_pages} pÃ¡gina(s) com checklist "
            "foram lidas corretamente â€” parte do layout pode ter mudado."
        )
    return issues


def get_unknown_items(pdf_path, mappings):
    doc = fitz.open(str(pdf_path))
    seen = {}
    for page_num in range(len(doc)):
        for item in extract_checklist(doc[page_num]):
            key = mapping_key(item["produto"], item["sku"])
            if key not in mappings and key not in seen:
                seen[key] = {
                    "produto": item["produto"], "sku": item["sku"], "key": key,
                    "file": pdf_path.name, "page": page_num + 1,
                }
    doc.close()
    return list(seen.values())


# â”€â”€ Holiday helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def get_br_holidays(year):
    result = {}
    for (month, day), name in BR_HOLIDAYS_FIXED.items():
        try:
            result[date(year, month, day).isoformat()] = name
        except ValueError:
            pass
    for (month, day), name in BR_HOLIDAYS_VARIABLE.get(year, {}).items():
        result[date(year, month, day).isoformat()] = name
    return result


def _make_period_filter(period, from_date=None, to_date=None):
    today     = date.today()
    yesterday = today - timedelta(days=1)
    def in_period(entry_date_str):
        try:
            ed = date.fromisoformat(entry_date_str)
        except Exception:
            return False
        if period == "today":     return ed == today
        if period == "yesterday": return ed == yesterday
        if period == "7d":        return (today - ed).days < 7
        if period == "15d":       return (today - ed).days < 15
        if period == "30d":       return (today - ed).days < 30
        if period == "custom" and from_date and to_date:
            return date.fromisoformat(from_date) <= ed <= date.fromisoformat(to_date)
        return True
    return in_period


def _date_range(period, filtered, from_date=None, to_date=None):
    today     = date.today()
    yesterday = today - timedelta(days=1)
    if period == "today":     return today, today
    if period == "yesterday": return yesterday, yesterday
    if period == "7d":        return today - timedelta(days=6), today
    if period == "15d":       return today - timedelta(days=14), today
    if period == "30d":       return today - timedelta(days=29), today
    if period == "custom" and from_date and to_date:
        return date.fromisoformat(from_date), date.fromisoformat(to_date)
    dates = [date.fromisoformat(e["date"]) for e in filtered if e.get("date")]
    if not dates:
        return today, today
    return min(dates), max(dates)


# â”€â”€ Cover page generator â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_SKIP_COVER = {"sem_produto", "sem_checklist", "pagina_em_branco",
               "ml_sem_produto", "ml_sem_checklist", "mlc2_sem_produto"}

def _cover_title(key):
    k = re.sub(r'^(?:mlc2_|ml_)', '', key)
    m = re.match(r'^([a-z]+)_(\d+)$', k)
    if m:
        return f"{m.group(1).upper()} {m.group(2)}"
    labels = {"roupas": "ROUPAS", "variacoes": "VARIAÃ‡Ã•ES"}
    return labels.get(k, k.upper())

def create_cover_page_bytes(title, pedidos, unidades, source="", label_date=""):
    """Portrait 4Ã—6 in cover page for thermal/Zebra label printer (black only)."""
    W, H  = 288, 432  # 4Ã—6 inches @ 72 dpi â€” portrait
    BLACK = (0, 0, 0)
    GRAY  = (0.55, 0.55, 0.55)
    doc   = fitz.open()
    page  = doc.new_page(width=W, height=H)

    font  = fitz.Font("hebo")   # Helvetica Bold (built-in)
    tw    = fitz.TextWriter(page.rect)

    def add_centered(text, baseline_y, fontsize):
        fs = fontsize
        while fs >= 8:
            if font.text_length(text, fs) <= W - 24:
                break
            fs -= 2
        x = (W - font.text_length(text, fs)) / 2
        tw.append((x, baseline_y), text, font=font, fontsize=fs)

    title_y = 160
    if source:
        add_centered(source, 72, 22)
        # draw thin separator line
        page.draw_line((40, 88), (W - 40, 88), color=GRAY, width=0.8)

    add_centered(title,                 title_y, 64)
    add_centered(f"{pedidos} PEDIDOS",  252,     48)
    add_centered(f"{unidades} UNIDADES",340,     48)

    if label_date:
        try:
            d = datetime.strptime(label_date, "%Y-%m-%d")
            display_date = d.strftime("%d/%m/%Y")
        except ValueError:
            display_date = label_date
        add_centered(display_date, 403, 22)

    tw.write_text(page, color=BLACK)

    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    buf.seek(0)
    return buf.read()


def _prepend_cover(writer, key, label_date=""):
    """Insert a cover page as the first page of writer, in-place."""
    if key in _SKIP_COVER:
        return
    pedidos = len(writer.pages)
    if pedidos == 0:
        return
    m      = re.match(r'^(?:mlc2_|ml_)?([a-z]+)_(\d+)$', key)
    upo    = int(m.group(2)) if m else 1
    source = "MERCADO LIVRE" if key.startswith("ml_") or key.startswith("mlc2_") else ""
    cover_bytes = create_cover_page_bytes(_cover_title(key), pedidos, pedidos * upo,
                                          source=source, label_date=label_date)
    cover_reader = PdfReader(io.BytesIO(cover_bytes))
    writer.insert_page(cover_reader.pages[0], index=0)


# â”€â”€ Core split/merge â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def compute_brand_totals(breakdown):
    brand_orders, brand_units = {}, {}
    for key, count in breakdown.items():
        if key in ("roupas", "roupas.pdf"):
            brand_orders["roupa"] = brand_orders.get("roupa", 0) + count
            brand_units["roupa"]  = brand_units.get("roupa", 0) + count
            continue
        if key.startswith("variacoes"):
            brand_orders["variacoes"] = brand_orders.get("variacoes", 0) + count
            brand_units["variacoes"]  = brand_units.get("variacoes", 0) + count
            continue
        m = re.match(r"^(.+)_(\d+)$", key)
        if not m:
            continue
        brand, upo = m.group(1), int(m.group(2))
        brand_orders[brand] = brand_orders.get(brand, 0) + count
        brand_units[brand]  = brand_units.get(brand, 0) + count * upo
    return brand_orders, brand_units


def split_pdfs_merged(filenames, mappings, label_date="", retirada_filenames=None,
                      skip_brs=None, keep_both_brs=None, detect_only=False,
                      declaracao_filenames=None):
    """
    Process multiple PDFs and merge pages with the same output key into one PDF each.
    Returns stats + output file list.

    skip_brs       : conjunto de BRs cujas páginas NÃO devem entrar no output (resolvidos como "não separar")
    keep_both_brs  : conjunto de BRs duplicados no batch que o usuário decidiu manter as 2 cópias
    detect_only    : se True, não grava PDFs em disco (só roda a detecção)
    declaracao_filenames : PDFs no formato Declaração de Conteúdo (ímpar=etiqueta, par=declaração)
    """
    retirada_set   = set(retirada_filenames or [])
    declaracao_set = set(declaracao_filenames or [])
    skip_brs      = set(skip_brs or [])
    keep_both_brs = set(keep_both_brs or [])
    seen_brs      = set()    # BRs já adicionados ao output neste batch (para dedupe)
    writers            = {}  # output_key -> PdfWriter
    page_results       = []
    store_names        = []
    store_items        = {}  # store -> {sku_key -> {sku, produto, categoria, kit_size, units}}
    store_label_counts = {}  # store -> number of label pages (1 page = 1 customer order)
    state_counts       = {}  # UF -> number of orders shipped there
    tracking_numbers     = {}  # tracking_number -> {filename, page}
    batch_tracking_dups  = {}  # tracking_number -> {file1, page1, file2, page2}
    skipped_dups         = 0   # quantas páginas foram puladas por duplicidade/skip
    total_pages_all    = 0
    blank_pages_all    = 0
    sem_check_all      = 0
    declaration_only_pages = 0  # páginas de declaração (não vão pro output)
    breakdown_all      = {}
    seen_sku_titles    = {}    # sku -> [titulos] (coletado no parse principal p/ detectar conflito sem re-parse)

    _t_start = _time.time()
    for filename in filenames:
        pdf_path = UPLOAD_DIR / filename
        doc      = fitz.open(str(pdf_path))
        reader   = PdfReader(str(pdf_path))
        n_pages  = len(doc)
        total_pages_all += n_pages

        store = extract_store_name(pdf_path)
        if store not in store_names:
            store_names.append(store)

        # ── Declaração de Conteúdo: ímpar=etiqueta, par=declaração (produto) ──
        if filename in declaracao_set:
            for i in range(0, n_pages - 1, 2):
                label_pg = doc[i]
                decl_pg  = doc[i + 1]
                declaration_only_pages += 1   # página de declaração não vai pro output
                items = extract_declaration_items(decl_pg)
                for _it in items:
                    _sk = _it["sku"].strip()
                    if _sk:
                        seen_sku_titles.setdefault(_sk, []).append(_it["produto"])
                if not items:
                    sem_check_all += 1
                    page_results.append({"file": filename, "page": i + 1,
                                          "items": [], "output": "sem_checklist"})
                    writers.setdefault("sem_checklist", PdfWriter()).add_page(reader.pages[i])
                    continue
                output_key = determine_output_pdf(items, mappings) or "__unknown__"
                page_results.append({
                    "file": filename, "page": i + 1, "output": output_key,
                    "items": [{"produto": x["produto"][:50], "sku": x["sku"],
                               "quantidade": x["quantidade"]} for x in items],
                })
                if output_key == "__unknown__":
                    continue
                page_tns = extract_tracking_numbers(label_pg)
                for tn in page_tns:
                    if tn in tracking_numbers:
                        prev = tracking_numbers[tn]
                        if prev["filename"] != filename and tn not in batch_tracking_dups:
                            batch_tracking_dups[tn] = {
                                "tracking": tn,
                                "file1": prev["filename"], "page1": prev["page"],
                                "file2": filename,         "page2": i + 1,
                            }
                skip_this = False
                for tn in page_tns:
                    if tn in skip_brs:
                        skip_this = True; break
                    if tn in seen_brs and tn not in keep_both_brs:
                        skip_this = True; break
                if skip_this:
                    skipped_dups += 1
                    page_results[-1]["output"] = "__skipped_dup__"
                    for tn in page_tns:
                        tracking_numbers.setdefault(tn, {"filename": filename, "page": i + 1,
                                                         "store": store, "items": []})
                    continue
                # adiciona SÓ a etiqueta (página i); a declaração nunca vai pro output
                writers.setdefault(output_key, PdfWriter()).add_page(reader.pages[i])
                breakdown_all[output_key] = breakdown_all.get(output_key, 0) + 1
                store_label_counts[store] = store_label_counts.get(store, 0) + 1
                uf = extract_destination_state(label_pg)
                if uf:
                    state_counts[uf] = state_counts.get(uf, 0) + 1
                page_items_info = []
                for item in items:
                    mk2 = mapping_key(item["produto"], item["sku"])
                    if mk2 in mappings:
                        info2 = mappings[mk2]
                        page_items_info.append({
                            "produto": item["produto"], "sku": item["sku"],
                            "quantidade": item["quantidade"],
                            "categoria": info2.get("categoria", "?"),
                            "kit_size": info2.get("kit_size", 1),
                        })
                for tn in page_tns:
                    seen_brs.add(tn)
                    if tn not in tracking_numbers or not tracking_numbers[tn].get("items"):
                        tracking_numbers[tn] = {"filename": filename, "page": i + 1,
                                                "store": store, "items": page_items_info}
                for item in items:
                    mk = mapping_key(item["produto"], item["sku"])
                    sku_key = item["sku"].strip()
                    if mk not in mappings:
                        continue
                    info = mappings[mk]; kit = info.get("kit_size", 1); qty = item["quantidade"]
                    store_items.setdefault(store, {})
                    if sku_key not in store_items[store]:
                        store_items[store][sku_key] = {
                            "sku": item["sku"], "produto": item["produto"],
                            "categoria": info["categoria"], "kit_size": kit, "units": 0,
                        }
                    store_items[store][sku_key]["units"] += kit * qty
            doc.close()
            continue

        _cl_fn = extract_checklist_combined if filename in retirada_set else extract_checklist
        for page_num in range(n_pages):
            fitz_page = doc[page_num]
            items     = _cl_fn(fitz_page)

            # Coleta sku+titulo p/ detecção de conflito (sem precisar re-parsear depois)
            for _it in items:
                _sk = _it["sku"].strip()
                if _sk:
                    seen_sku_titles.setdefault(_sk, []).append(_it["produto"])

            if not items:
                text = fitz_page.get_text("text").strip()
                if len(text) < 20:
                    output_key = "pagina_em_branco"
                    blank_pages_all += 1
                else:
                    output_key = "sem_checklist"
                    sem_check_all  += 1
                page_results.append({
                    "file": filename, "page": page_num + 1,
                    "items": [], "output": output_key,
                })
            else:
                output_key = determine_output_pdf(items, mappings) or "__unknown__"
                page_results.append({
                    "file": filename, "page": page_num + 1,
                    "output": output_key,
                    "items": [{"produto": i["produto"][:50], "sku": i["sku"],
                               "quantidade": i["quantidade"]} for i in items],
                })

            if output_key in ("pagina_em_branco", "__unknown__"):
                continue

            # ── Tracking numbers desta página (antes de adicionar ao output) ──
            page_tns = extract_tracking_numbers(fitz_page) if items else []

            # Detecção de duplicidade dentro do batch (sempre roda, pra alimentar o popup)
            for tn in page_tns:
                if tn in tracking_numbers:
                    prev = tracking_numbers[tn]
                    if prev["filename"] != filename and tn not in batch_tracking_dups:
                        batch_tracking_dups[tn] = {
                            "tracking": tn,
                            "file1": prev["filename"], "page1": prev["page"],
                            "file2": filename,         "page2": page_num + 1,
                        }
                elif page_tns:
                    pass  # registrado abaixo

            # ── Decide se pula esta página (skip explícito ou dedupe de batch) ──
            skip_this = False
            for tn in page_tns:
                if tn in skip_brs:
                    skip_this = True; break
                if tn in seen_brs and tn not in keep_both_brs:
                    skip_this = True; break   # 2ª cópia do mesmo BR no batch → mantém só 1
            if skip_this:
                skipped_dups += 1
                page_results[-1]["output"] = "__skipped_dup__"
                # registra o BR como visto (pra contabilizar conflito), mas não adiciona página
                for tn in page_tns:
                    tracking_numbers.setdefault(tn, {"filename": filename, "page": page_num + 1,
                                                     "store": store, "items": []})
                continue

            writers.setdefault(output_key, PdfWriter()).add_page(reader.pages[page_num])
            if output_key not in ("sem_checklist",):
                breakdown_all[output_key] = breakdown_all.get(output_key, 0) + 1

            # Collect per-store item stats (one pass per page)
            if items:
                store_label_counts[store] = store_label_counts.get(store, 0) + 1
                uf = extract_destination_state(fitz_page)
                if uf:
                    state_counts[uf] = state_counts.get(uf, 0) + 1
                page_items_info = []
                for item in items:
                    mk2 = mapping_key(item["produto"], item["sku"])
                    if mk2 in mappings:
                        info2 = mappings[mk2]
                        page_items_info.append({
                            "produto":   item["produto"],
                            "sku":       item["sku"],
                            "quantidade": item["quantidade"],
                            "categoria": info2.get("categoria", "?"),
                            "kit_size":  info2.get("kit_size", 1),
                        })
                for tn in page_tns:
                    seen_brs.add(tn)
                    if tn not in tracking_numbers or not tracking_numbers[tn].get("items"):
                        tracking_numbers[tn] = {
                            "filename": filename,
                            "page":     page_num + 1,
                            "store":    store,
                            "items":    page_items_info,
                        }
                for item in items:
                    mk      = mapping_key(item["produto"], item["sku"])
                    sku_key = item["sku"].strip()
                    if mk not in mappings:
                        continue
                    info = mappings[mk]
                    kit  = info.get("kit_size", 1)
                    qty  = item["quantidade"]
                    store_items.setdefault(store, {})
                    if sku_key not in store_items[store]:
                        store_items[store][sku_key] = {
                            "sku":       item["sku"],
                            "produto":   item["produto"],
                            "categoria": info["categoria"],
                            "kit_size":  kit,
                            "units":     0,
                        }
                    store_items[store][sku_key]["units"] += kit * qty

        doc.close()

    _t_parse = _time.time()
    print(f"[PERF] parse de {len(filenames)} PDF(s) / {total_pages_all} pags: {_t_parse-_t_start:.1f}s", flush=True)

    # Write output PDFs (with cover page prepended) — pulado no modo detect_only
    output_files       = []
    output_page_counts = {}
    for key, writer in writers.items():
        cover_offset = 0 if key in _SKIP_COVER else 1
        if not detect_only:
            _prepend_cover(writer, key, label_date=label_date)
            out_path = OUTPUT_DIR / f"{key}.pdf"
            with open(out_path, "wb") as f:
                writer.write(f)
        output_files.append(f"{key}.pdf")
        # subtract cover page from counts used for verification
        npages = len(writer.pages) - (cover_offset if not detect_only else 0)
        output_page_counts[f"{key}.pdf"] = npages
    _t_write = _time.time()
    print(f"[PERF] escrita de {len(writers)} PDF(s) de saida: {_t_write-_t_parse:.1f}s (detect_only={detect_only})", flush=True)

    # label_pages efetivo desconta páginas puladas por duplicidade/skip e as de declaração
    label_pages  = total_pages_all - blank_pages_all - sem_check_all - skipped_dups - declaration_only_pages
    output_total = sum(c for k, c in output_page_counts.items()
                       if not k.startswith("sem_checklist"))
    verified = output_total == label_pages

    brand_orders, brand_units = compute_brand_totals(breakdown_all)
    brand_display = get_brand_display()
    brand_summary = sorted(
        [{"brand": b, "display": brand_display.get(b, b),
          "orders": brand_orders[b], "units": brand_units[b]}
         for b in brand_orders],
        key=lambda x: (
            STANDARD_ORDER.index(x["brand"]) if x["brand"] in STANDARD_ORDER else len(STANDARD_ORDER),
            x["brand"]
        )
    )

    # Security: check tracking duplicates vs history log
    tracking_log      = load_tracking_log()
    tracking_conflicts = []
    for tn, info in tracking_numbers.items():
        if tn in tracking_log:
            tracking_conflicts.append({
                "tracking": tn,
                "previous_date": tracking_log[tn].get("date", "?"),
                "filename": info["filename"],
                "page": info["page"],
            })

    # Security: detect SKU+title mismatches (usa títulos já coletados no parse — sem re-parsear)
    _mapped_skus = {}
    for _mk, _info in mappings.items():
        if "|" in _mk:
            _mapped_skus.setdefault(_mk.split("|", 1)[0], []).append((_mk, _info))
    _conf = {}
    for _sku, _titles in seen_sku_titles.items():
        if _sku not in _mapped_skus:
            continue
        for _titulo in _titles:
            _mk = _sku + "|" + _normalize_titulo(_titulo)
            if _mk in mappings or _sku in _conf:
                continue
            _ex = _mapped_skus[_sku]
            _conf[_sku] = {
                "sku": _sku,
                "titulo_encontrado": re.sub(r"\s+", " ", _titulo).strip(),
                "titulo_salvo": _ex[0][1].get("titulo", _ex[0][0].split("|", 1)[-1]),
                "categoria_atual": _ex[0][1].get("categoria", "?"),
                "kit_size": _ex[0][1].get("kit_size", 1),
            }
    sku_conflicts = list(_conf.values())
    print(f"[PERF] deteccao conflito: {_time.time()-_t_write:.2f}s | TOTAL split: {_time.time()-_t_start:.1f}s", flush=True)

    stats = {
        "total_pages":        total_pages_all,
        "label_pages":        label_pages,
        "blank_pages":        blank_pages_all,
        "sem_checklist":      sem_check_all,
        "store_names":        store_names,
        "store_items":        store_items,
        "store_label_counts": store_label_counts,
        "state_counts":       state_counts,
        "breakdown":          breakdown_all,
        "output_page_counts": output_page_counts,
        "brand_summary":      brand_summary,
        "tracking_numbers":   list(tracking_numbers.keys()),
        "tracking_data":      tracking_numbers,
        "skipped_dups":       skipped_dups,
        "verification": {
            "ok":           verified,
            "label_pages":  label_pages,
            "output_total": output_total,
        },
    }

    security_alerts = {
        "tracking_conflicts":      tracking_conflicts,
        "sku_conflicts":           sku_conflicts,
        "batch_tracking_dups":     list(batch_tracking_dups.values()),
    }

    return {"output_files": output_files, "page_results": page_results,
            "stats": stats, "errors": [], "security_alerts": security_alerts}


# â”€â”€ Mercado Livre â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

ML_STORE_NAME = "REPZILLA ML"


def is_ml_label_page(page):
    return "Cidade de destino" in page.get_text("text")


def extract_ml_label_ids(page):
    """Return (pack_id, venda_id) from an ML label page. Either may be None."""
    text     = page.get_text("text")
    pack_id  = None
    venda_id = None
    m = re.search(r'Pack\s+ID:\s*(\d+)', text, re.IGNORECASE)
    if m:
        pack_id = m.group(1)
    m = re.search(r'Venda:\s*(\d+)', text, re.IGNORECASE)
    if m:
        venda_id = m.group(1)
    return pack_id, venda_id


def extract_ml_label_state(page):
    """Extract destination state from 'Cidade de destino' line."""
    m = re.search(r'Cidade\s+de\s+destino\s*[:\-]?\s*([^\n]+)',
                  page.get_text("text"), re.IGNORECASE)
    if not m:
        return None
    dest = m.group(1).strip()
    for sep in ('/', ',', ' - '):
        if sep in dest:
            tail = dest.rsplit(sep, 1)[-1].strip()
            uf = tail[:2].upper()
            if uf in BR_STATES:
                return uf
    dest_u = dest.upper()
    for name, uf in BR_STATE_NAMES:
        if name in dest_u:
            return uf
    for tok in dest_u.split():
        if len(tok) == 2 and tok in BR_STATES:
            return tok
    return None


_ML_PACKID_RE = re.compile(r'^Pack\s+ID:\s*(\d+)', re.IGNORECASE)
_ML_VENDA_RE  = re.compile(r'^Venda:\s*(\d+)', re.IGNORECASE)
_ML_SKU_RE    = re.compile(r'^SKU:\s*(\S+)', re.IGNORECASE)
_ML_QTY_RE    = re.compile(r'^Quantidade:\s*(\d+)', re.IGNORECASE)


def parse_ml_product_pages(doc):
    """Parse product identification pages in an ML PDF.

    Each page has:
      - Blocks: Pack ID / Venda / customer name / SKU / Quantidade (in order)
      - Footer: "Despachem as suas vendas..." followed by one product title per entry

    Titles at the bottom are paired positionally with entries in block order.
    Keyed by both Pack ID and Venda ID.
    """
    product_map = {}
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text("text")
        if "Checklist" in text or "REMETENTE" in text:
            continue
        if is_ml_label_page(page):
            continue
        if "SKU:" not in text:
            continue
        lines = [l.strip() for l in text.split('\n') if l.strip()]

        # Split at "Despachem" footer: titles live after it
        despachem_idx = next(
            (i for i, l in enumerate(lines) if l.lower().startswith("despachem")), None
        )
        block_lines  = lines[:despachem_idx] if despachem_idx is not None else lines
        page_titles  = lines[despachem_idx + 1:] if despachem_idx is not None else []

        # Parse blocks in order, collecting (pack_id, venda_id, sku, quantidade)
        entries = []
        i = 0
        while i < len(block_lines):
            m_pack  = _ML_PACKID_RE.match(block_lines[i])
            m_venda = _ML_VENDA_RE.match(block_lines[i])
            if not (m_pack or m_venda):
                i += 1; continue

            pack_id  = m_pack.group(1)  if m_pack  else None
            venda_id = m_venda.group(1) if m_venda else None
            sku = None; quantidade = 1
            j = i + 1
            while j < min(i + 15, len(block_lines)):
                if _ML_PACKID_RE.match(block_lines[j]):
                    break
                mv = _ML_VENDA_RE.match(block_lines[j])
                if mv:
                    if venda_id:
                        break
                    venda_id = mv.group(1)
                    j += 1; continue
                ms = _ML_SKU_RE.match(block_lines[j])
                mq = _ML_QTY_RE.match(block_lines[j])
                if ms:
                    sku = ms.group(1).strip()
                elif mq:
                    try: quantidade = int(mq.group(1))
                    except ValueError: pass
                j += 1

            if sku:
                entries.append({"pack_id": pack_id, "venda_id": venda_id,
                                 "sku": sku, "quantidade": quantidade})
            i = j

        # Pair each entry with its positional title from the footer section
        for idx, e in enumerate(entries):
            titulo = page_titles[idx] if idx < len(page_titles) else e["sku"]
            item = {"sku": e["sku"], "produto": titulo, "quantidade": e["quantidade"], "page": page_num + 1}
            if e["pack_id"]:
                product_map.setdefault(e["pack_id"], []).append(item)
            if e["venda_id"]:
                product_map.setdefault(e["venda_id"], []).append(item)

    return product_map


def get_unknown_items_ml(pdf_path, mappings):
    doc = fitz.open(str(pdf_path))
    product_map = parse_ml_product_pages(doc)
    doc.close()
    seen: dict = {}
    for items in product_map.values():
        for item in items:
            k = mapping_key(item["produto"], item["sku"])
            if k not in mappings and k not in seen:
                seen[k] = {
                    "produto": item["produto"], "sku": item["sku"], "key": k,
                    "file": pdf_path.name, "page": item.get("page", 1),
                }
    return list(seen.values())


def split_ml_pdfs(filenames, mappings, label_date=""):
    """Process ML PDFs â€” same return structure as split_pdfs_merged."""
    writers              = {}
    page_results         = []
    store_items          = {}
    store_label_counts   = {}
    state_counts         = {}
    tracking_numbers     = {}
    batch_tracking_dups  = {}
    total_pages_all      = 0
    unmatched_all        = 0
    breakdown_all        = {}
    store                = ML_STORE_NAME

    for filename in filenames:
        pdf_path = UPLOAD_DIR / filename
        doc      = fitz.open(str(pdf_path))
        reader   = PdfReader(str(pdf_path))
        total_pages_all += len(doc)
        product_map = parse_ml_product_pages(doc)

        for page_num in range(len(doc)):
            fp    = doc[page_num]
            if not is_ml_label_page(fp):
                page_results.append({
                    "file": filename, "page": page_num + 1,
                    "items": [], "output": "product_info_page",
                })
                continue
            pack_id, venda_id = extract_ml_label_ids(fp)
            items = product_map.get(pack_id, []) if pack_id else []
            if not items and venda_id:
                items = product_map.get(venda_id, [])
            if not items:
                unmatched_all += 1
                writers.setdefault("ml_sem_produto", PdfWriter()).add_page(reader.pages[page_num])
                page_results.append({
                    "file": filename, "page": page_num + 1,
                    "items": [], "output": "ml_sem_produto",
                })
                continue
            checklist_items = [
                {"produto": it["produto"], "sku": it["sku"], "quantidade": it["quantidade"]}
                for it in items
            ]
            raw_key = determine_output_pdf(checklist_items, mappings)
            if raw_key is None:
                page_results.append({
                    "file": filename, "page": page_num + 1,
                    "output": "__unknown__", "items": checklist_items,
                })
                continue
            ml_key = "ml_" + raw_key
            writers.setdefault(ml_key, PdfWriter()).add_page(reader.pages[page_num])
            breakdown_all[ml_key] = breakdown_all.get(ml_key, 0) + 1
            store_label_counts[store] = store_label_counts.get(store, 0) + 1
            uf = extract_ml_label_state(fp)
            if uf:
                state_counts[uf] = state_counts.get(uf, 0) + 1
            for tn in extract_tracking_numbers(fp):
                if tn not in tracking_numbers:
                    tracking_numbers[tn] = {"filename": filename, "page": page_num + 1}
                else:
                    prev = tracking_numbers[tn]
                    if prev["filename"] != filename and tn not in batch_tracking_dups:
                        batch_tracking_dups[tn] = {
                            "tracking": tn,
                            "file1": prev["filename"], "page1": prev["page"],
                            "file2": filename,         "page2": page_num + 1,
                        }
            page_results.append({
                "file": filename, "page": page_num + 1, "output": ml_key,
                "items": [{"produto": it["produto"][:50], "sku": it["sku"],
                            "quantidade": it["quantidade"]} for it in items],
            })
            for item in checklist_items:
                mk      = mapping_key(item["produto"], item["sku"])
                sku_key = item["sku"].strip()
                if mk not in mappings:
                    continue
                info = mappings[mk]
                kit  = info.get("kit_size", 1)
                store_items.setdefault(store, {})
                if sku_key not in store_items[store]:
                    store_items[store][sku_key] = {
                        "sku":       item["sku"],
                        "produto":   item["produto"],
                        "categoria": info["categoria"],
                        "kit_size":  kit,
                        "units":     0,
                    }
                store_items[store][sku_key]["units"] += kit * item["quantidade"]
        doc.close()

    output_files       = []
    output_page_counts = {}
    for key, writer in writers.items():
        _prepend_cover(writer, key, label_date=label_date)
        out_path = OUTPUT_DIR / f"{key}.pdf"
        with open(out_path, "wb") as f:
            writer.write(f)
        output_files.append(f"{key}.pdf")
        cover_offset = 0 if key in _SKIP_COVER else 1
        output_page_counts[f"{key}.pdf"] = len(writer.pages) - cover_offset

    label_pages  = store_label_counts.get(store, 0)
    output_total = sum(c for k, c in output_page_counts.items()
                       if k != "ml_sem_produto.pdf")
    verified     = (output_total == label_pages)

    stripped = {k.replace("ml_", "", 1): v for k, v in breakdown_all.items()}
    brand_orders, brand_units = compute_brand_totals(stripped)
    brand_display = get_brand_display()
    brand_summary = sorted(
        [{"brand": b, "display": brand_display.get(b, b),
          "orders": brand_orders[b], "units": brand_units[b]}
         for b in brand_orders],
        key=lambda x: (
            STANDARD_ORDER.index(x["brand"]) if x["brand"] in STANDARD_ORDER else len(STANDARD_ORDER),
            x["brand"]
        )
    )

    tracking_log       = load_tracking_log()
    tracking_conflicts = []
    for tn, info in tracking_numbers.items():
        if tn in tracking_log:
            tracking_conflicts.append({
                "tracking": tn,
                "previous_date": tracking_log[tn].get("date", "?"),
                "filename": info["filename"],
                "page": info["page"],
            })

    stats = {
        "total_pages":        total_pages_all,
        "label_pages":        label_pages,
        "blank_pages":        0,
        "sem_checklist":      unmatched_all,
        "store_names":        [store],
        "store_items":        store_items,
        "store_label_counts": store_label_counts,
        "state_counts":       state_counts,
        "breakdown":          breakdown_all,
        "output_page_counts": output_page_counts,
        "brand_summary":      brand_summary,
        "tracking_numbers":   list(tracking_numbers.keys()),
        "tracking_data":      tracking_numbers,
        "verification":       {
            "ok":           verified,
            "label_pages":  label_pages,
            "output_total": output_total,
        },
    }

    security_alerts = {
        "tracking_conflicts":  tracking_conflicts,
        "sku_conflicts":       [],
        "batch_tracking_dups": list(batch_tracking_dups.values()),
    }

    return {"output_files": output_files, "page_results": page_results,
            "stats": stats, "errors": [], "security_alerts": security_alerts}


# â”€â”€ Auth â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def hash_pin(pin):
    return hashlib.sha256(pin.encode()).hexdigest()

# â”€â”€ Auditoria â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def load_auditoria():
    if AUDITORIA_FILE.exists():
        with open(AUDITORIA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_auditoria(hist):
    with open(AUDITORIA_FILE, "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)

def registrar_aud(acao, descricao):
    usuario = session.pop("audit_user", None) or session.get("user", "?")
    hist = load_auditoria()
    hist.insert(0, {
        "ts":        datetime.now().isoformat(),
        "usuario":   usuario,
        "acao":      acao,
        "descricao": descricao,
    })
    save_auditoria(hist[:1000])

def _valida_pin_session():
    """Verifica se um PIN foi confirmado nos Ãºltimos 30s e consome o token."""
    ts = session.pop("pin_ok_ts", 0)
    return (_time.time() - ts) <= 30

def load_users():
    if USERS_FILE.exists():
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_users(users):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

@app.before_request
def require_login():
    public = {"login", "setup_pin", "logout", "static"}
    if request.endpoint in public:
        return
    if "user" not in session:
        return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    users = load_users()
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").upper().strip()
        pin = request.form.get("pin", "").strip()
        if username not in users:
            error = "UsuÃ¡rio invÃ¡lido."
        elif users[username]["pin"] is None:
            return redirect(url_for("setup_pin", username=username))
        elif users[username]["pin"] != hash_pin(pin):
            error = "PIN incorreto."
        else:
            session.permanent = True
            session["user"] = username
            return redirect(url_for("index"))
    return render_template("login.html", users=list(users.keys()), error=error)

@app.route("/setup-pin", methods=["GET", "POST"])
def setup_pin():
    username = (request.args.get("username") or request.form.get("username", "")).upper().strip()
    users = load_users()
    error = None
    if username not in users:
        return redirect(url_for("login"))
    if request.method == "POST":
        pin = request.form.get("pin", "").strip()
        pin_confirm = request.form.get("pin_confirm", "").strip()
        if len(pin) != 4 or not pin.isdigit():
            error = "O PIN deve ter exatamente 4 nÃºmeros."
        elif pin != pin_confirm:
            error = "Os PINs nÃ£o coincidem. Tente novamente."
        else:
            users[username]["pin"] = hash_pin(pin)
            save_users(users)
            session.permanent = True
            session["user"] = username
            return redirect(url_for("index"))
    return render_template("setup_pin.html", username=username, error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/api/confirmar-pin", methods=["POST"])
def api_confirmar_pin():
    data         = request.json or {}
    pin          = data.get("pin", "").strip()
    usuario_alvo = (data.get("username") or session.get("user", "")).upper().strip()
    users        = load_users()
    if not usuario_alvo or usuario_alvo not in users:
        return jsonify({"error": "UsuÃ¡rio invÃ¡lido"}), 403
    if users[usuario_alvo]["pin"] != hash_pin(pin):
        registrar_aud("pin_falhou", f"PIN incorreto para usuÃ¡rio {usuario_alvo}")
        return jsonify({"error": "PIN incorreto"}), 403
    session["pin_ok_ts"]   = _time.time()
    session["audit_user"]  = usuario_alvo
    return jsonify({"ok": True})

@app.route("/api/auditoria", methods=["GET"])
def api_auditoria():
    limit = int(request.args.get("limit", 200))
    return jsonify(load_auditoria()[:limit])

# â”€â”€ Routes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/")
def index():
    users = load_users()
    resp = make_response(render_template("index.html",
        current_user=session.get("user", ""),
        all_users=list(users.keys())
    ))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.route("/upload", methods=["POST"])
def upload():
    files      = request.files.getlist("pdfs")
    retirada   = request.args.get("retirada") == "1"
    declaracao = request.args.get("declaracao") == "1"
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    mappings    = load_mappings()
    saved       = []
    all_unknown = {}
    store_names = []
    batch_skus  = {}
    sku_stores  = {}
    if declaracao:
        _unknown_fn = get_unknown_items_declaracao
    elif retirada:
        _unknown_fn = get_unknown_items_combined
    else:
        _unknown_fn = get_unknown_items

    for file in files:
        if not file.filename.lower().endswith(".pdf"):
            continue
        pdf_path = UPLOAD_DIR / file.filename
        file.save(str(pdf_path))
        saved.append(file.filename)
        s = extract_store_name(pdf_path)
        if s not in store_names:
            store_names.append(s)
        for u in _unknown_fn(pdf_path, mappings):
            all_unknown.setdefault(u["key"], u)
        doc = fitz.open(str(pdf_path))
        if declaracao:
            for page_num in range(1, len(doc), 2):   # páginas de declaração
                for item in extract_declaration_items(doc[page_num]):
                    base = item["sku"].strip()
                    batch_skus.setdefault(base, set()).add(_normalize_titulo(item["produto"]))
                    sku_stores.setdefault(base, set()).add(s)
        else:
            _extract_fn = extract_checklist_combined if retirada else extract_checklist
            for page_num in range(len(doc)):
                for item in _extract_fn(doc[page_num]):
                    base = item["sku"].strip()
                    norm = _normalize_titulo(item["produto"])
                    batch_skus.setdefault(base, set()).add(norm)
                    sku_stores.setdefault(base, set()).add(s)
        doc.close()

    if not saved:
        return jsonify({"error": "Nenhum PDF vÃ¡lido encontrado"}), 400

    # Base SKUs with 2+ titles in batch where at least one title is NOT yet mapped
    dup_skus = set()
    for base, titles in batch_skus.items():
        if len(titles) > 1:
            if any(f"{base}|{t}" not in mappings for t in titles):
                dup_skus.add(base)
    has_batch_sku_dups = bool(dup_skus)
    # Only return stores for duplicate SKUs with unmapped entries
    sku_stores_out = {base: sorted(sku_stores[base]) for base in dup_skus}

    format_warnings = {}
    if not declaracao:
        for fname in saved:
            issues = check_label_format(UPLOAD_DIR / fname)
            if issues:
                format_warnings[fname] = issues

    return jsonify({
        "filenames":           saved,
        "unknown":             list(all_unknown.values()),
        "format_warnings":     format_warnings,
        "store_names":         store_names,
        "has_batch_sku_dups":  has_batch_sku_dups,
        "sku_stores":          sku_stores_out,
    })


@app.route("/save-mappings", methods=["POST"])
def save_mappings_route():
    data = request.json
    mappings = load_mappings()
    mappings.update(data.get("mappings", {}))
    save_mappings(mappings)
    return jsonify({"ok": True})


@app.route("/process", methods=["POST"])
def process():
    data               = request.json
    filenames          = data.get("filenames", [])
    retirada_filenames = data.get("retirada_filenames", [])
    declaracao_filenames = data.get("declaracao_filenames", [])
    label_date         = data.get("label_date", "")
    skip_brs           = data.get("skip_brs", [])
    keep_both_brs      = data.get("keep_both_brs", [])
    resolved           = data.get("resolved", False)   # usuário já passou pela tela de problemas
    force              = data.get("force", False)       # "prosseguir mesmo com erros"
    all_fnames         = filenames + retirada_filenames + declaracao_filenames
    if not all_fnames:
        return jsonify({"error": "Nenhum arquivo especificado"}), 400

    mappings    = load_mappings()
    all_unknown = {}
    retirada_set   = set(retirada_filenames)
    declaracao_set = set(declaracao_filenames)
    for filename in all_fnames:
        pdf_path = UPLOAD_DIR / filename
        if not pdf_path.exists():
            return jsonify({"error": f"Arquivo nÃ£o encontrado: {filename}"}), 404
        if filename in declaracao_set:
            _fn = get_unknown_items_declaracao
        elif filename in retirada_set:
            _fn = get_unknown_items_combined
        else:
            _fn = get_unknown_items
        for u in _fn(pdf_path, mappings):
            all_unknown.setdefault(u["key"], u)

    if all_unknown:
        return jsonify({"error": "Ainda hÃ¡ produtos nÃ£o classificados",
                        "unknown": list(all_unknown.values())}), 400

    # Passada única: gera os PDFs aplicando skips/keep-both (vazios na 1ª chamada)
    result = split_pdfs_merged(all_fnames, mappings, label_date=label_date,
                               retirada_filenames=retirada_filenames,
                               declaracao_filenames=declaracao_filenames,
                               skip_brs=skip_brs, keep_both_brs=keep_both_brs)

    # Trava: na 1ª chamada (sem resolução/force), se houver conflitos → exige resolução.
    # Os PDFs gerados aqui são descartados e regerados corretos na chamada de resolução.
    if not resolved and not force:
        alerts = result.get("security_alerts", {})
        has_conflicts = bool(alerts.get("tracking_conflicts") or
                             alerts.get("batch_tracking_dups") or
                             alerts.get("sku_conflicts"))
        if has_conflicts:
            return jsonify({
                "needs_resolution": True,
                "security_alerts":  alerts,
                "stats":            {"label_pages": result["stats"]["label_pages"]},
            })

    return jsonify(result)


@app.route("/download/<filename>")
def download(filename):
    today = date.today().strftime("%d.%m.%y")
    stem  = filename.replace(".pdf", "").replace("_", " ")
    return send_from_directory(str(OUTPUT_DIR), filename, as_attachment=True,
                               download_name=f"{stem} {today}.pdf")


def _pdf_with_cover(pdf_path, key, label_date, saved_opc, breakdown):
    """Return BytesIO of the PDF with cover page regenerated for the given date/counts."""
    if key in _SKIP_COVER:
        with open(pdf_path, "rb") as f:
            return io.BytesIO(f.read())
    pedidos = saved_opc.get(key + ".pdf", breakdown.get(key, 0))
    m       = re.match(r'^(?:ml_)?([a-z]+)_(\d+)$', key)
    upo     = int(m.group(2)) if m else 1
    source  = "MERCADO LIVRE" if key.startswith("ml_") else ""
    cover_bytes  = create_cover_page_bytes(_cover_title(key), pedidos, pedidos * upo,
                                           source=source, label_date=label_date)
    reader = PdfReader(pdf_path)
    writer = PdfWriter()
    for i, pg in enumerate(reader.pages):
        if i > 0:
            writer.add_page(pg)
    writer.insert_page(PdfReader(io.BytesIO(cover_bytes)).pages[0], index=0)
    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    return buf


def _archive_pdf_path(entry, filename):
    """Retorna o caminho do PDF no arquivo da expediÃ§Ã£o, ou None se nÃ£o existir."""
    archive_id = entry.get("archive_dir")
    if archive_id:
        p = ARCHIVE_DIR / archive_id / filename
        if p.exists():
            return p
    return None


@app.route("/history/download/<int:idx>/<path:filename>")
def history_download(idx, filename):
    history = load_history()
    if not (0 <= idx < len(history)):
        return jsonify({"error": "ExpediÃ§Ã£o nÃ£o encontrada"}), 404
    if not filename.endswith(".pdf") or "/" in filename or "\\" in filename:
        return jsonify({"error": "Arquivo invÃ¡lido"}), 400
    entry      = history[idx]
    label_date = entry.get("date", "")
    breakdown  = entry.get("totals", {}).get("breakdown", {})
    saved_opc  = entry.get("output_page_counts", {})
    key        = filename.replace(".pdf", "")

    # Usa arquivo da expediÃ§Ã£o se disponÃ­vel, senÃ£o cai no disco atual
    pdf_path = _archive_pdf_path(entry, filename) or (OUTPUT_DIR / filename)
    if not pdf_path.exists():
        return jsonify({"error": "Arquivo nÃ£o encontrado no disco"}), 404

    buf = _pdf_with_cover(pdf_path, key, label_date, saved_opc, breakdown)
    try:
        d = datetime.strptime(label_date, "%Y-%m-%d")
        date_str = d.strftime("%d.%m.%y")
    except Exception:
        date_str = date.today().strftime("%d.%m.%y")
    stem = filename.replace(".pdf", "").replace("_", " ")
    return send_file(buf, as_attachment=True,
                     download_name=f"{stem} {date_str}.pdf",
                     mimetype="application/pdf")


@app.route("/history/download-all/<int:idx>", methods=["POST"])
def history_download_all(idx):
    history = load_history()
    if not (0 <= idx < len(history)):
        return jsonify({"error": "ExpediÃ§Ã£o nÃ£o encontrada"}), 404
    entry      = history[idx]
    label_date = entry.get("date", "")
    breakdown  = entry.get("totals", {}).get("breakdown", {})
    saved_opc  = entry.get("output_page_counts", {})
    filenames  = request.json.get("filenames", [])
    try:
        d = datetime.strptime(label_date, "%Y-%m-%d")
        date_str = d.strftime("%d.%m.%y")
    except Exception:
        date_str = date.today().strftime("%d.%m.%y")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname in filenames:
            path = _archive_pdf_path(entry, fname) or (OUTPUT_DIR / fname)
            if not path.exists():
                continue
            key  = fname.replace(".pdf", "")
            stem = fname.replace(".pdf", "").replace("_", " ")
            pdf_buf = _pdf_with_cover(path, key, label_date, saved_opc, breakdown)
            zf.writestr(f"{stem} {date_str}.pdf", pdf_buf.read())
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"etiquetas {date_str}.zip",
                     mimetype="application/zip")


@app.route("/download-all", methods=["POST"])
def download_all():
    data      = request.json
    filenames = data.get("filenames", [])
    today     = date.today().strftime("%d.%m.%y")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname in filenames:
            path = OUTPUT_DIR / fname
            if path.exists():
                stem = fname.replace(".pdf", "").replace("_", " ")
                zf.write(str(path), f"{stem} {today}.pdf")
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"etiquetas {today}.zip",
                     mimetype="application/zip")


@app.route("/mappings", methods=["GET"])
def get_mappings():
    return jsonify(load_mappings())


@app.route("/mappings/sku-duplicates", methods=["GET"])
def mappings_sku_duplicates():
    """Return groups of mapping entries that share the same base SKU."""
    mappings = load_mappings()
    by_sku = {}
    for mk, info in mappings.items():
        base = mk.split("|", 1)[0] if "|" in mk else mk
        by_sku.setdefault(base, []).append({
            "key":      mk,
            "titulo":   info.get("titulo", mk.split("|", 1)[-1] if "|" in mk else mk),
            "categoria": info.get("categoria", "?"),
            "kit_size": info.get("kit_size", 1),
        })
    duplicates = {sku: entries for sku, entries in by_sku.items() if len(entries) > 1}
    return jsonify(duplicates)


@app.route("/pdf-page/<path:filename>/<int:page_num>")
def pdf_page_image(filename, page_num):
    """Return a single PDF page rendered as PNG for the classify preview."""
    import re
    from flask import Response
    if not re.match(r'^[\w\-. ]+\.pdf$', filename, re.IGNORECASE):
        return jsonify({"error": "invalid filename"}), 400
    pdf_path = UPLOAD_DIR / filename
    if not pdf_path.exists():
        return jsonify({"error": "not found"}), 404
    try:
        doc = fitz.open(str(pdf_path))
        if page_num < 1 or page_num > len(doc):
            doc.close()
            return jsonify({"error": "page out of range"}), 400
        page = doc[page_num - 1]
        pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
        img_bytes = pix.tobytes("png")
        doc.close()
        return Response(img_bytes, mimetype="image/png")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/mappings/delete", methods=["POST"])
def delete_mapping():
    data     = request.json
    mappings = load_mappings()
    mappings.pop(data.get("key", ""), None)
    save_mappings(mappings)
    return jsonify({"ok": True})


@app.route("/brands", methods=["GET"])
def get_brands_route():
    return jsonify(get_brand_display())


@app.route("/brands/save", methods=["POST"])
def save_brand_route():
    data    = request.json
    slug    = data.get("slug", "").strip()
    display = data.get("display", "").strip()
    if not slug or not display:
        return jsonify({"error": "slug e display sÃ£o obrigatÃ³rios"}), 400
    brands = load_brands()
    brands[slug] = display
    save_brands(brands)
    return jsonify({"ok": True})


@app.route("/history/add", methods=["POST"])
def history_add():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data        = request.json
    history     = load_history()
    entry_date  = data.get("date") or date.today().isoformat()
    expedition_id = datetime.now().strftime("%Y%m%d-%H%M%S")
    entry       = {
        "id":                 expedition_id,
        "timestamp":          datetime.now().isoformat(),
        "date":               entry_date,
        "store_items":         data.get("store_items", {}),
        "store_label_counts":  data.get("store_label_counts", {}),
        "state_counts":        data.get("state_counts", {}),
        "output_page_counts":  data.get("output_page_counts", {}),
        "totals": {
            "label_pages": data.get("label_pages", 0),
            "breakdown":   data.get("breakdown", {}),
        },
    }

    # Arquivar PDFs desta expediÃ§Ã£o para downloads histÃ³ricos corretos
    breakdown = data.get("breakdown", {})
    opc       = data.get("output_page_counts", {})
    archive_exp_dir = ARCHIVE_DIR / expedition_id
    archive_exp_dir.mkdir(exist_ok=True)
    archived = []
    for fname in list(opc.keys()):
        src = OUTPUT_DIR / fname
        if src.exists():
            shutil.copy2(str(src), str(archive_exp_dir / fname))
            archived.append(fname)
    if archived:
        entry["archive_dir"] = expedition_id

    history.append(entry)
    save_history(history)

    # Save new tracking numbers to log (skip ones already handled by /security-resolve)
    new_tracking = data.get("tracking_numbers", [])
    if new_tracking:
        tracking_log = load_tracking_log()
        for tn in new_tracking:
            if tn not in tracking_log:
                tracking_log[tn] = {"date": entry_date}
        save_tracking_log(tracking_log)

    # Save enriched tracking index (tracking_number -> product info)
    tracking_data = data.get("tracking_data", {})
    if tracking_data:
        index = load_tracking_index()
        for tn, info in tracking_data.items():
            index[tn] = {
                "date":         entry_date,
                "expedition_id": entry["id"],
                "store":        info.get("store", ""),
                "filename":     info.get("filename", ""),
                "page":         info.get("page", 0),
                "items":        info.get("items", []),
            }
        save_tracking_index(index)

    total = sum(data.get("store_label_counts", {}).values()) or data.get("label_pages", 0)
    registrar_aud("lacrar", f"ExpediÃ§Ã£o {entry_date} â€” {total} etiquetas lacradas")
    return jsonify({"ok": True, "id": entry["id"]})


@app.route("/security-resolve", methods=["POST"])
def security_resolve():
    """Resolve security alerts: tracking duplicates and SKU+title conflicts."""
    data    = request.json
    today   = date.today().isoformat()

    # Tracking resolutions: {tracking_number: "delete_previous" | "keep_both" | "ignore_today"}
    tracking_actions = data.get("tracking_actions", {})
    tracking_log     = load_tracking_log()
    for tn, action in tracking_actions.items():
        if action == "delete_previous":
            # Replace old entry with today's date
            tracking_log[tn] = {"date": today}
        elif action == "keep_both":
            # Save today alongside the existing entry
            existing = tracking_log.get(tn, {})
            tracking_log[tn] = {
                "date": today,
                "also_seen": existing.get("date"),
            }
        # "ignore_today": don't touch the log for this number

    save_tracking_log(tracking_log)

    # SKU conflict resolutions: [{sku, titulo_encontrado, categoria, kit_size}]
    sku_actions = data.get("sku_actions", [])
    if sku_actions:
        mappings = load_mappings()
        for action in sku_actions:
            sku    = action.get("sku", "")
            titulo = action.get("titulo_encontrado", "") or action.get("titulo", "")
            cat    = action.get("categoria")
            kit    = action.get("kit_size", 1)
            if sku and titulo and cat:
                mk = mapping_key(titulo, sku)
                mappings[mk] = {"categoria": cat, "kit_size": kit, "titulo": titulo}
        save_mappings(mappings)

    return jsonify({"ok": True})


@app.route("/history/metrics", methods=["POST"])
def history_metrics():
    data      = request.json
    stores    = data.get("stores", [])
    period    = data.get("period", "all")
    from_date = data.get("from_date")
    to_date   = data.get("to_date")
    history   = load_history()
    today     = date.today()
    yesterday = today - timedelta(days=1)

    def in_period(entry_date_str):
        try:
            ed = date.fromisoformat(entry_date_str)
        except Exception:
            return False
        if period == "today":     return ed == today
        if period == "yesterday": return ed == yesterday
        if period == "7d":        return (today - ed).days < 7
        if period == "15d":       return (today - ed).days < 15
        if period == "30d":       return (today - ed).days < 30
        if period == "custom" and from_date and to_date:
            return date.fromisoformat(from_date) <= ed <= date.fromisoformat(to_date)
        return True

    filtered = [e for e in history if in_period(e.get("date", ""))]

    store_totals   = {}
    product_totals = {}

    for entry in filtered:
        slc = entry.get("store_label_counts", {})
        # Fallback for old entries: approximate label count from product units (best effort)
        if not slc:
            slc = {s: entry.get("totals", {}).get("label_pages", 0)
                   for s in entry.get("store_items", {})}

        for store_name, items in entry.get("store_items", {}).items():
            if stores and store_name not in stores:
                continue
            label_count = slc.get(store_name, 0)
            if store_name not in store_totals:
                store_totals[store_name] = {"orders": 0, "units": 0, "produtos": {}}
            store_totals[store_name]["orders"] += label_count
            for sku_key, item in items.items():
                units = item.get("units", 0)
                store_totals[store_name]["units"] += units
                st_prod = store_totals[store_name]["produtos"]
                if sku_key not in st_prod:
                    st_prod[sku_key] = {**item, "units": 0}
                st_prod[sku_key]["units"] += units
                if sku_key not in product_totals:
                    product_totals[sku_key] = {**item, "units": 0}
                product_totals[sku_key]["units"] += units

    # Top products grouped by categoria
    brand_display  = get_brand_display()
    cat_totals     = {}
    for item in product_totals.values():
        cat = item.get("categoria", "other")
        if cat not in cat_totals:
            display = "Roupas" if cat == "roupa" else brand_display.get(cat, cat.capitalize())
            cat_totals[cat] = {"categoria": cat, "display": display, "units": 0}
        cat_totals[cat]["units"] += item.get("units", 0)
    top_products = sorted(cat_totals.values(), key=lambda x: x["units"], reverse=True)
    # Kit vs avulso: label counts where those skus appeared
    kit_units  = sum(v["units"] for v in product_totals.values() if v.get("kit_size", 1) > 1)
    solo_units = sum(v["units"] for v in product_totals.values() if v.get("kit_size", 1) == 1)
    all_stores = sorted({s for e in history for s in e.get("store_items", {})})

    total_orders = sum(s["orders"] for s in store_totals.values())
    total_units  = sum(s["units"]  for s in store_totals.values())

    return jsonify({
        "store_totals":  store_totals,
        "top_products":  top_products,
        "kit_vs_solo":   {"kit": kit_units, "solo": solo_units},
        "all_stores":    all_stores,
        "total_entries": len(filtered),
        "summary": {
            "total_orders": total_orders,
            "total_units":  total_units,
        },
    })


@app.route("/history/states", methods=["POST"])
def history_states():
    data      = request.json
    period    = data.get("period", "all")
    from_date = data.get("from_date")
    to_date   = data.get("to_date")
    history   = load_history()
    in_period = _make_period_filter(period, from_date, to_date)
    filtered  = [e for e in history if in_period(e.get("date", ""))]
    state_totals = {}
    for entry in filtered:
        for uf, count in entry.get("state_counts", {}).items():
            state_totals[uf] = state_totals.get(uf, 0) + count
    return jsonify({"state_totals": state_totals})


@app.route("/history/daily", methods=["POST"])
def history_daily():
    data      = request.json
    stores    = data.get("stores", [])
    period    = data.get("period", "all")
    from_date = data.get("from_date")
    to_date   = data.get("to_date")
    history   = load_history()

    in_period = _make_period_filter(period, from_date, to_date)
    filtered  = [e for e in history if in_period(e.get("date", ""))]
    if not filtered:
        return jsonify({"days": [], "holidays": {}})

    start_d, end_d = _date_range(period, filtered, from_date, to_date)
    years = set()
    cur = start_d
    while cur <= end_d:
        years.add(cur.year); cur += timedelta(days=1)
    holidays = {}
    for y in years:
        holidays.update(get_br_holidays(y))

    WEEKDAYS = ["Segunda","TerÃ§a","Quarta","Quinta","Sexta","SÃ¡bado","Domingo"]
    daily = {}
    cur = start_d
    while cur <= end_d:
        iso = cur.isoformat()
        daily[iso] = {
            "date": iso, "weekday": WEEKDAYS[cur.weekday()],
            "holiday": holidays.get(iso),
            "orders": 0, "units": 0,
            "by_categoria": {}, "by_store": {},
        }
        cur += timedelta(days=1)

    for entry in filtered:
        d_str = entry.get("date", "")
        if d_str not in daily:
            continue
        slc = entry.get("store_label_counts", {})
        day = daily[d_str]
        for store_name, items in entry.get("store_items", {}).items():
            if stores and store_name not in stores:
                continue
            label_count = slc.get(store_name, 0)
            day["orders"] += label_count
            day["by_store"][store_name] = day["by_store"].get(store_name, 0) + label_count
            for sku_key, item in items.items():
                units = item.get("units", 0)
                cat   = item.get("categoria", "other")
                day["units"] += units
                day["by_categoria"][cat] = day["by_categoria"].get(cat, 0) + units

    return jsonify({"days": list(daily.values()), "holidays": holidays})


@app.route("/history/products-list", methods=["GET"])
def history_products_list():
    history  = load_history()
    products = {}
    for entry in history:
        for store_name, items in entry.get("store_items", {}).items():
            for sku_key, item in items.items():
                if sku_key not in products:
                    products[sku_key] = {
                        "sku_key":   sku_key,
                        "sku":       item.get("sku", sku_key),
                        "produto":   item.get("produto", sku_key),
                        "categoria": item.get("categoria", ""),
                        "kit_size":  item.get("kit_size", 1),
                    }
    return jsonify(sorted(products.values(), key=lambda x: x["produto"]))


@app.route("/history/product-metrics", methods=["POST"])
def history_product_metrics():
    data      = request.json
    categoria = data.get("categoria", "")
    stores    = data.get("stores", [])
    period    = data.get("period", "all")
    from_date = data.get("from_date")
    to_date   = data.get("to_date")
    history   = load_history()

    in_period = _make_period_filter(period, from_date, to_date)
    filtered  = [e for e in history if in_period(e.get("date", ""))]
    if not filtered or not categoria:
        return jsonify({"days":[],"store_totals":{},"holidays":{},"category_info":{}})

    brand_display = get_brand_display()
    cat_display   = "Roupas" if categoria == "roupa" else brand_display.get(categoria, categoria.capitalize())
    category_info = {"categoria": categoria, "display": cat_display}

    start_d, end_d = _date_range(period, filtered, from_date, to_date)
    years = set()
    cur = start_d
    while cur <= end_d:
        years.add(cur.year); cur += timedelta(days=1)
    holidays = {}
    for y in years:
        holidays.update(get_br_holidays(y))

    WEEKDAYS = ["Segunda","TerÃ§a","Quarta","Quinta","Sexta","SÃ¡bado","Domingo"]
    daily = {}
    cur = start_d
    while cur <= end_d:
        iso = cur.isoformat()
        daily[iso] = {
            "date": iso, "weekday": WEEKDAYS[cur.weekday()],
            "holiday": holidays.get(iso),
            "total_units": 0, "by_store": {},
        }
        cur += timedelta(days=1)

    store_totals = {}
    for entry in filtered:
        d_str = entry.get("date", "")
        if d_str not in daily:
            continue
        for store_name, items in entry.get("store_items", {}).items():
            if stores and store_name not in stores:
                continue
            for sku_key, item in items.items():
                if item.get("categoria") != categoria:
                    continue
                units = item.get("units", 0)
                daily[d_str]["total_units"] += units
                daily[d_str]["by_store"][store_name] = \
                    daily[d_str]["by_store"].get(store_name, 0) + units
                store_totals[store_name] = store_totals.get(store_name, 0) + units

    return jsonify({
        "days":          list(daily.values()),
        "store_totals":  store_totals,
        "holidays":      holidays,
        "category_info": category_info,
    })


@app.route("/history/entries", methods=["GET"])
def history_entries_list():
    history      = load_history()
    estoque      = load_estoque()
    brand_display = get_brand_display()
    entries = []
    for i, entry in enumerate(history):
        slc = entry.get("store_label_counts", {})
        total_orders = sum(slc.values()) if slc else entry.get("totals", {}).get("label_pages", 0)
        stores = list(slc.keys()) or list(entry.get("store_items", {}).keys())
        breakdown_raw = entry.get("totals", {}).get("breakdown", {})
        breakdown_items = []
        for grupo, qtd in breakdown_raw.items():
            if not isinstance(qtd, (int, float)) or qtd <= 0:
                continue
            display = (estoque.get(grupo, {}).get("display")
                       or brand_display.get(grupo, grupo))
            breakdown_items.append({"grupo": grupo, "display": display, "quantidade": int(qtd)})
        entries.append({
            "index":           i,
            "date":            entry.get("date", "?"),
            "id":              entry.get("id", str(i)),
            "stores":          stores,
            "store_orders":    {s: slc.get(s, 0) for s in stores},
            "total_orders":    total_orders,
            "breakdown_items": breakdown_items,
        })
    # Sort by date descending so newest is first regardless of insertion order
    entries.sort(key=lambda e: e["date"], reverse=True)
    return jsonify({"entries": entries})


@app.route("/history/delete", methods=["POST"])
def history_delete():
    if not _valida_pin_session():
        return jsonify({"ok": False, "error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data          = request.json
    idx           = data.get("index")
    restore_stock = data.get("restore_stock", False)
    history       = load_history()
    if idx is None or not (0 <= idx < len(history)):
        return jsonify({"ok": False, "error": "index invÃ¡lido"}), 400
    entry      = history[idx]
    entry_date = entry.get("date", "?")

    if restore_stock:
        breakdown    = entry.get("totals", {}).get("breakdown", {})
        estoque      = load_estoque()
        hist_est     = load_hist_estoque()
        brand_display = get_brand_display()
        movimentos   = []
        for grupo, qtd in breakdown.items():
            if not isinstance(qtd, (int, float)) or qtd <= 0:
                continue
            if grupo not in estoque:
                continue
            estoque[grupo]["quantidade"] += int(qtd)
            display = estoque[grupo].get("display") or brand_display.get(grupo, grupo)
            movimentos.append({"grupo": grupo, "display": display, "quantidade": int(qtd)})
        if movimentos:
            hist_est.append({
                "id":            secrets.token_hex(8),
                "timestamp":     datetime.now().isoformat(timespec="seconds"),
                "usuario":       session.get("user", "?"),
                "tipo":          "estorno_historico",
                "itens":         movimentos,
                "justificativa": f"Estorno â€” expediÃ§Ã£o {entry_date} excluÃ­da",
                "contexto":      entry.get("id", ""),
            })
            save_estoque(estoque)
            save_hist_estoque(hist_est)

    history.pop(idx)
    save_history(history)
    registrar_aud("del_hist", f"ExpediÃ§Ã£o de {entry_date} excluÃ­da do histÃ³rico")
    return jsonify({"ok": True})


@app.route("/history/delete-store", methods=["POST"])
def history_delete_store():
    if not _valida_pin_session():
        return jsonify({"ok": False, "error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data    = request.json
    idx     = data.get("index")
    store   = data.get("store", "")
    history = load_history()
    if idx is None or not (0 <= idx < len(history)):
        return jsonify({"ok": False, "error": "index invÃ¡lido"}), 400
    entry = history[idx]
    entry_date = entry.get("date", "?")
    entry.get("store_items", {}).pop(store, None)
    entry.get("store_label_counts", {}).pop(store, None)
    remaining = (list(entry.get("store_items", {}).keys())
                 or list(entry.get("store_label_counts", {}).keys()))
    if not remaining:
        history.pop(idx)
    else:
        new_lp = sum(entry.get("store_label_counts", {}).values())
        entry.setdefault("totals", {})["label_pages"] = new_lp
        history[idx] = entry
    save_history(history)
    registrar_aud("del_hist_loja", f"Loja '{store}' removida da expediÃ§Ã£o {entry_date}")
    return jsonify({"ok": True})


@app.route("/history/reopen/<int:idx>", methods=["GET"])
def history_reopen(idx):
    history = load_history()
    if not (0 <= idx < len(history)):
        return jsonify({"error": "ExpediÃ§Ã£o nÃ£o encontrada"}), 404
    entry    = history[idx]
    breakdown = entry.get("totals", {}).get("breakdown", {})
    label_pages = entry.get("totals", {}).get("label_pages", 0)
    store_items         = entry.get("store_items", {})
    store_label_counts  = entry.get("store_label_counts", {})

    # Use saved output_page_counts from history (set at lacrar time, covers excluded)
    saved_opc = entry.get("output_page_counts", {})

    # Check which output files still exist on disk (for download links only)
    output_files       = []
    output_page_counts = {}
    for key in breakdown:
        fname = key + ".pdf"
        fpath = OUTPUT_DIR / fname
        if fpath.exists():
            output_files.append(fname)
        if fname in saved_opc:
            output_page_counts[fname] = saved_opc[fname]
        elif fpath.exists():
            # Fallback for old history entries that don't have saved counts
            try:
                doc = fitz.open(str(fpath))
                output_page_counts[fname] = max(0, len(doc) - 1)
                doc.close()
            except Exception:
                output_page_counts[fname] = 0

    brand_orders, brand_units = compute_brand_totals(breakdown)
    brand_display = get_brand_display()
    brand_summary = sorted(
        [{"brand": b, "display": brand_display.get(b, b),
          "orders": brand_orders[b], "units": brand_units[b]}
         for b in brand_orders],
        key=lambda x: (
            STANDARD_ORDER.index(x["brand"]) if x["brand"] in STANDARD_ORDER else len(STANDARD_ORDER),
            x["brand"]
        )
    )

    store_names  = list(store_label_counts.keys()) or list(store_items.keys())
    saved_total  = sum(output_page_counts.values()) if output_page_counts else label_pages

    stats = {
        "label_pages":        label_pages,
        "total_pages":        saved_total,
        "blank_pages":        0,
        "sem_checklist":      0,
        "store_names":        store_names,
        "breakdown":          breakdown,
        "output_page_counts": output_page_counts,
        "brand_summary":      brand_summary,
        "is_history_reopen":  True,
        "verification": {
            "ok":           True,
            "label_pages":  label_pages,
            "output_total": label_pages,
        },
    }
    return jsonify({
        "stats":         stats,
        "output_files":  output_files,
        "files_on_disk": len(output_files),
        "date":          entry.get("date", ""),
    })


@app.route("/metrics/export", methods=["POST"])
def metrics_export():
    d              = request.json
    period_label   = d.get("period_label", "PerÃ­odo selecionado")
    summary        = d.get("summary", {})
    store_totals   = d.get("store_totals", {})
    all_prods      = d.get("all_products_data", [])
    state_totals   = d.get("state_totals", {})

    total_orders = summary.get("total_orders", 0)
    store_rows = sorted(store_totals.items(), key=lambda x: x[1].get("orders", 0), reverse=True)
    state_rows = sorted(state_totals.items(), key=lambda x: x[1], reverse=True)
    total_states = sum(v for _, v in state_rows)

    def pct(n, t): return f"{round(n/t*100)}%" if t else "â€”"

    rows_store = "".join(
        f"<tr><td>{s}</td><td class='n'>{v.get('orders',0)}</td>"
        f"<td class='n'>{pct(v.get('orders',0),total_orders)}</td>"
        f"<td class='n'>{v.get('units',0)}</td></tr>"
        for s, v in store_rows
    )
    prod_blocks = ""
    for p in all_prods:
        cat  = p.get("cat", "")
        disp = p.get("display", cat)
        st   = p.get("store_totals", {})
        if not st: continue
        tot  = p.get("total_units", 0)
        ul   = "peÃ§as" if cat == "roupa" else "potes"
        rows = "".join(
            f"<tr><td>{s}</td><td class='n'>{v}</td><td class='n'>{pct(v,tot)}</td></tr>"
            for s, v in sorted(st.items(), key=lambda x: x[1], reverse=True)
        )
        prod_blocks += f"""
        <div class="pb">
          <div class="pn">{disp} <span class="badge">{ul}</span></div>
          <table><tr><th>Loja</th><th class='n'>{ul.capitalize()}</th><th class='n'>%</th></tr>{rows}</table>
        </div>"""

    state_block = ""
    if state_rows:
        state_rows_html = "".join(
            f"<tr><td>{uf}</td><td class='n'>{cnt}</td><td class='n'>{pct(cnt,total_states)}</td></tr>"
            for uf, cnt in state_rows
        )
        state_block = f"""
        <div class="sec">Estados</div>
        <table><tr><th>UF</th><th class='n'>Pedidos</th><th class='n'>%</th></tr>{state_rows_html}</table>"""

    html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>REPZILLA â€” MÃ©tricas</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;color:#1a1a2e;padding:32px;font-size:13px;max-width:860px;margin:auto}}
h1{{font-size:24px;font-weight:900;letter-spacing:-1px;margin-bottom:3px}}
.sub{{font-size:11px;color:#888;margin-bottom:24px}}
.sec{{font-size:9px;font-weight:800;text-transform:uppercase;letter-spacing:1px;color:#8888a8;margin:22px 0 8px;border-bottom:2px solid #e2e2ec;padding-bottom:4px}}
.stat-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:0}}
.sb{{border:1px solid #e2e2ec;border-radius:8px;padding:14px;text-align:center}}
.sv{{font-size:28px;font-weight:900;line-height:1;margin-bottom:3px}}
.sl{{font-size:10px;color:#8888a8}}
table{{width:100%;border-collapse:collapse;margin-bottom:0}}
th{{font-size:9px;font-weight:800;text-transform:uppercase;letter-spacing:.5px;color:#8888a8;text-align:left;padding:6px 8px;border-bottom:2px solid #e2e2ec}}
td{{padding:7px 8px;border-bottom:1px solid #f0f0f8;font-size:13px}}
.n{{text-align:right;font-weight:700;font-variant-numeric:tabular-nums}}
.pb{{margin-top:14px}}
.pn{{font-size:14px;font-weight:700;margin-bottom:6px}}
.badge{{font-size:9px;font-weight:800;text-transform:uppercase;padding:2px 7px;border-radius:4px;background:#f0f0f8;color:#8888a8;margin-left:6px;vertical-align:middle}}
.print-btn{{margin-bottom:20px}}
.print-btn button{{padding:9px 22px;background:#6c5fff;color:#fff;border:none;border-radius:7px;font-size:13px;font-weight:700;cursor:pointer}}
@media print{{.print-btn{{display:none}}.pb{{page-break-inside:avoid}}}}
</style></head><body>
<div class="print-btn"><button onclick="window.print()">ðŸ–¨ï¸ Salvar como PDF</button></div>
<h1>REPZILLA</h1>
<p class="sub">RelatÃ³rio de MÃ©tricas &nbsp;Â·&nbsp; {period_label}</p>
<div class="sec">Resumo</div>
<div class="stat-grid">
  <div class="sb"><div class="sv">{summary.get('total_orders',0)}</div><div class="sl">pedidos totais</div></div>
  <div class="sb"><div class="sv">{summary.get('total_units',0)}</div><div class="sl">unidades totais</div></div>
  <div class="sb"><div class="sv">{len(store_totals)}</div><div class="sl">lojas ativas</div></div>
  <div class="sb"><div class="sv">{summary.get('total_entries',0)}</div><div class="sl">expediÃ§Ãµes</div></div>
</div>
<div class="sec">Pedidos por Loja</div>
<table><tr><th>Loja</th><th class='n'>Pedidos</th><th class='n'>%</th><th class='n'>Unidades</th></tr>{rows_store}</table>
<div class="sec">AnÃ¡lise por Produto</div>
{prod_blocks}
{state_block}
</body></html>"""
    from flask import Response
    return Response(html, mimetype="text/html")


@app.route("/ml/upload", methods=["POST"])
def ml_upload():
    files = request.files.getlist("pdfs")
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    mappings    = load_mappings()
    saved       = []
    all_unknown = {}
    for file in files:
        if not file.filename.lower().endswith(".pdf"):
            continue
        pdf_path = UPLOAD_DIR / file.filename
        file.save(str(pdf_path))
        saved.append(file.filename)
        for u in get_unknown_items_ml(pdf_path, mappings):
            all_unknown.setdefault(u["key"], u)
    if not saved:
        return jsonify({"error": "Nenhum PDF vÃ¡lido encontrado"}), 400
    return jsonify({"filenames": saved, "unknown": list(all_unknown.values())})


@app.route("/ml/process", methods=["POST"])
def ml_process():
    data       = request.json
    filenames  = data.get("filenames", [])
    label_date = data.get("label_date", "")
    if not filenames:
        return jsonify({"error": "Nenhum arquivo especificado"}), 400
    mappings    = load_mappings()
    all_unknown = {}
    for filename in filenames:
        pdf_path = UPLOAD_DIR / filename
        if not pdf_path.exists():
            return jsonify({"error": f"Arquivo nÃ£o encontrado: {filename}"}), 404
        for u in get_unknown_items_ml(pdf_path, mappings):
            all_unknown.setdefault(u["key"], u)
    if all_unknown:
        return jsonify({"error": "Ainda hÃ¡ produtos nÃ£o classificados",
                        "unknown": list(all_unknown.values())}), 400
    return jsonify(split_ml_pdfs(filenames, mappings, label_date=label_date))


@app.route("/ml/download-all", methods=["POST"])
def ml_download_all():
    data      = request.json
    filenames = data.get("filenames", [])
    today     = date.today().strftime("%d.%m.%y")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname in filenames:
            path = OUTPUT_DIR / fname
            if path.exists():
                stem = fname.replace(".pdf", "").replace("ml_", "").replace("_", " ")
                zf.write(str(path), f"ML {stem} {today}.pdf")
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"ml_etiquetas_{today}.zip",
                     mimetype="application/zip")


@app.route("/ml/correios", methods=["POST"])
def ml_correios():
    """Extract label + product-checklist pages from a Correios-format ML PDF.
    Skips DANFE pages; keeps everything else in original order."""
    f = request.files.get("file")
    if not f or not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Envie um arquivo PDF"}), 400

    raw  = f.read()
    src  = fitz.open(stream=raw, filetype="pdf")
    out  = fitz.open()
    labels = 0

    for page in src:
        text = page.get_text("text")
        if "DANFE" in text:
            continue
        out.insert_pdf(src, from_page=page.number, to_page=page.number)
        if "SHP:" in text:
            labels += 1

    if len(out) == 0:
        src.close(); out.close()
        return jsonify({"error": "Nenhuma pÃ¡gina vÃ¡lida encontrada. Verifique o arquivo."}), 400

    today    = date.today().strftime("%d.%m.%y")
    fname    = f"ml_correios_{today}.pdf"
    out_path = OUTPUT_DIR / fname
    out.save(str(out_path))
    total = len(out)
    out.close(); src.close()

    return jsonify({"filename": fname, "labels": labels, "total": total})


# â”€â”€ ML Correios 2 â€” SeparaÃ§Ã£o â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_CORREIOS_TRACKING_RE = re.compile(r'\b([A-Z]{2}\d{9}BR)\b')
_CORREIOS_QTY_RE      = re.compile(r'^Quantidade:\s*(\d+)', re.IGNORECASE)


def parse_correios_product_pages(doc):
    """Parse product identification pages from a clean Correios-format ML PDF.
    Returns {tracking_code: [{"produto": title, "sku": "", "quantidade": N}]}
    """
    product_map = {}
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text("text")
        if "SHP:" in text or "DANFE" in text:
            continue
        if "Quantidade:" not in text:
            continue
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        despachem_idx = next(
            (i for i, l in enumerate(lines) if l.lower().startswith("despachem")), None
        )
        block_lines = lines[:despachem_idx] if despachem_idx is not None else lines
        page_titles = lines[despachem_idx + 1:] if despachem_idx is not None else []
        entries = []
        i = 0
        while i < len(block_lines):
            if not re.fullmatch(r'[A-Z]{2}\d{9}BR', block_lines[i]):
                i += 1; continue
            tracking_code = block_lines[i]
            quantidade = 1
            j = i + 1
            while j < len(block_lines):
                if re.fullmatch(r'[A-Z]{2}\d{9}BR', block_lines[j]):
                    break
                mq = _CORREIOS_QTY_RE.match(block_lines[j])
                if mq:
                    try: quantidade = int(mq.group(1))
                    except ValueError: pass
                j += 1
            entries.append({"tracking_code": tracking_code, "quantidade": quantidade})
            i = j
        for idx, e in enumerate(entries):
            titulo = page_titles[idx] if idx < len(page_titles) else "Produto desconhecido"
            item = {"sku": "", "produto": titulo, "quantidade": e["quantidade"]}
            product_map.setdefault(e["tracking_code"], []).append(item)
    return product_map


def extract_correios_label_tracking(page):
    """Return the postal tracking code from a Correios label page, or None."""
    m = _CORREIOS_TRACKING_RE.search(page.get_text("text"))
    return m.group(1) if m else None


def get_unknown_items_correios(pdf_path, mappings):
    doc = fitz.open(str(pdf_path))
    product_map = parse_correios_product_pages(doc)
    doc.close()
    seen = {}
    for items in product_map.values():
        for item in items:
            k = mapping_key(item["produto"], item["sku"])
            if k not in mappings and k not in seen:
                seen[k] = {"produto": item["produto"], "sku": item["sku"], "key": k,
                           "file": pdf_path.name, "page": 1}
    return list(seen.values())


def split_ml_correios2_pdf(filenames, mappings, label_date=""):
    """Process clean Correios PDFs â€” same return structure as split_ml_pdfs."""
    writers            = {}
    page_results       = []
    store_items        = {}
    store_label_counts = {}
    tracking_numbers   = {}
    total_pages_all    = 0
    unmatched_all      = 0
    breakdown_all      = {}
    store              = ML_STORE_NAME

    for filename in filenames:
        pdf_path = UPLOAD_DIR / filename
        doc      = fitz.open(str(pdf_path))
        reader   = PdfReader(str(pdf_path))
        total_pages_all += len(doc)
        product_map = parse_correios_product_pages(doc)

        for page_num in range(len(doc)):
            fp   = doc[page_num]
            text = fp.get_text("text")
            if "SHP:" not in text:
                page_results.append({
                    "file": filename, "page": page_num + 1,
                    "items": [], "output": "product_info_page",
                })
                continue
            tracking = extract_correios_label_tracking(fp)
            items = product_map.get(tracking, []) if tracking else []
            if not items:
                unmatched_all += 1
                writers.setdefault("mlc2_sem_produto", PdfWriter()).add_page(reader.pages[page_num])
                page_results.append({
                    "file": filename, "page": page_num + 1,
                    "items": [], "output": "mlc2_sem_produto",
                })
                continue
            checklist_items = [
                {"produto": it["produto"], "sku": it["sku"], "quantidade": it["quantidade"]}
                for it in items
            ]
            raw_key = determine_output_pdf(checklist_items, mappings)
            if raw_key is None:
                page_results.append({
                    "file": filename, "page": page_num + 1,
                    "output": "__unknown__", "items": checklist_items,
                })
                continue
            mlc2_key = "mlc2_" + raw_key
            writers.setdefault(mlc2_key, PdfWriter()).add_page(reader.pages[page_num])
            breakdown_all[mlc2_key] = breakdown_all.get(mlc2_key, 0) + 1
            store_label_counts[store] = store_label_counts.get(store, 0) + 1
            if tracking and tracking not in tracking_numbers:
                tracking_numbers[tracking] = {"filename": filename, "page": page_num + 1}
            page_results.append({
                "file": filename, "page": page_num + 1, "output": mlc2_key,
                "items": [{"produto": it["produto"][:50], "sku": it["sku"],
                            "quantidade": it["quantidade"]} for it in items],
            })
            for item in checklist_items:
                mk  = mapping_key(item["produto"], item["sku"])
                if mk not in mappings:
                    continue
                info = mappings[mk]
                kit  = info.get("kit_size", 1)
                store_items.setdefault(store, {})
                if mk not in store_items[store]:
                    store_items[store][mk] = {
                        "sku":       item["sku"],
                        "produto":   item["produto"],
                        "categoria": info["categoria"],
                        "kit_size":  kit,
                        "units":     0,
                    }
                store_items[store][mk]["units"] += kit * item["quantidade"]
        doc.close()

    output_files       = []
    output_page_counts = {}
    for key, writer in writers.items():
        _prepend_cover(writer, key, label_date=label_date)
        out_path = OUTPUT_DIR / f"{key}.pdf"
        with open(out_path, "wb") as f:
            writer.write(f)
        output_files.append(f"{key}.pdf")
        cover_offset = 0 if key in _SKIP_COVER else 1
        output_page_counts[f"{key}.pdf"] = len(writer.pages) - cover_offset

    label_pages  = store_label_counts.get(store, 0)
    output_total = sum(c for k, c in output_page_counts.items()
                       if k != "mlc2_sem_produto.pdf")
    verified     = (output_total == label_pages)

    stripped = {k.replace("mlc2_", "", 1): v for k, v in breakdown_all.items()}
    brand_orders, brand_units = compute_brand_totals(stripped)
    brand_display = get_brand_display()
    brand_summary = sorted(
        [{"brand": b, "display": brand_display.get(b, b),
          "orders": brand_orders[b], "units": brand_units[b]}
         for b in brand_orders],
        key=lambda x: (
            STANDARD_ORDER.index(x["brand"]) if x["brand"] in STANDARD_ORDER else len(STANDARD_ORDER),
            x["brand"]
        )
    )

    tracking_log       = load_tracking_log()
    tracking_conflicts = []
    for tn, info in tracking_numbers.items():
        if tn in tracking_log:
            tracking_conflicts.append({
                "tracking": tn,
                "previous_date": tracking_log[tn].get("date", "?"),
                "filename": info["filename"],
                "page": info["page"],
            })

    stats = {
        "total_pages":        total_pages_all,
        "label_pages":        label_pages,
        "blank_pages":        0,
        "sem_checklist":      unmatched_all,
        "store_names":        [store],
        "store_items":        store_items,
        "store_label_counts": store_label_counts,
        "state_counts":       {},
        "breakdown":          breakdown_all,
        "output_page_counts": output_page_counts,
        "brand_summary":      brand_summary,
        "tracking_numbers":   list(tracking_numbers.keys()),
        "tracking_data":      tracking_numbers,
        "verification":       {
            "ok":           verified,
            "label_pages":  label_pages,
            "output_total": output_total,
        },
    }

    security_alerts = {
        "tracking_conflicts": tracking_conflicts,
        "sku_conflicts":      [],
    }

    return {"output_files": output_files, "page_results": page_results,
            "stats": stats, "errors": [], "security_alerts": security_alerts}


@app.route("/ml/correios2/upload", methods=["POST"])
def ml_correios2_upload():
    files = request.files.getlist("pdfs")
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    mappings    = load_mappings()
    saved       = []
    all_unknown = {}
    for file in files:
        if not file.filename.lower().endswith(".pdf"):
            continue
        pdf_path = UPLOAD_DIR / file.filename
        file.save(str(pdf_path))
        saved.append(file.filename)
        for u in get_unknown_items_correios(pdf_path, mappings):
            all_unknown.setdefault(u["key"], u)
    if not saved:
        return jsonify({"error": "Nenhum PDF vÃ¡lido encontrado"}), 400
    return jsonify({"filenames": saved, "unknown": list(all_unknown.values())})


@app.route("/ml/correios2/process", methods=["POST"])
def ml_correios2_process():
    data       = request.json
    filenames  = data.get("filenames", [])
    label_date = data.get("label_date", "")
    if not filenames:
        return jsonify({"error": "Nenhum arquivo especificado"}), 400
    mappings    = load_mappings()
    all_unknown = {}
    for filename in filenames:
        pdf_path = UPLOAD_DIR / filename
        if not pdf_path.exists():
            return jsonify({"error": f"Arquivo nÃ£o encontrado: {filename}"}), 404
        for u in get_unknown_items_correios(pdf_path, mappings):
            all_unknown.setdefault(u["key"], u)
    if all_unknown:
        return jsonify({"error": "Ainda hÃ¡ produtos nÃ£o classificados",
                        "unknown": list(all_unknown.values())}), 400
    return jsonify(split_ml_correios2_pdf(filenames, mappings, label_date=label_date))


@app.route("/ml/correios2/download-all", methods=["POST"])
def ml_correios2_download_all():
    data      = request.json
    filenames = data.get("filenames", [])
    today     = date.today().strftime("%d.%m.%y")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname in filenames:
            path = OUTPUT_DIR / fname
            if path.exists():
                stem = fname.replace(".pdf", "").replace("mlc2_", "").replace("_", " ")
                zf.write(str(path), f"ML Correios {stem} {today}.pdf")
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"ml_correios_etiquetas_{today}.zip",
                     mimetype="application/zip")


@app.route("/order-locator", methods=["POST"])
def order_locator():
    """Search tracking numbers in the index. Returns found/not-found per number."""
    data     = request.json
    numbers  = [n.strip().upper() for n in data.get("numbers", []) if n.strip()]
    index    = load_tracking_index()
    found    = {}
    not_found = []
    for tn in numbers:
        if tn in index:
            entry = index[tn]
            # Check if source PDF still exists on disk
            src_available = (UPLOAD_DIR / entry.get("filename", "")).exists() if entry.get("filename") else False
            found[tn] = {**entry, "src_available": src_available}
        else:
            not_found.append(tn)
    return jsonify({"found": found, "not_found": not_found})


@app.route("/order-locator/download", methods=["POST"])
def order_locator_download():
    """Extract label pages for given tracking numbers and return as merged PDF."""
    data    = request.json
    numbers = [n.strip().upper() for n in data.get("numbers", []) if n.strip()]
    index   = load_tracking_index()

    writer  = PdfWriter()
    missing = []
    for tn in numbers:
        if tn not in index:
            missing.append(tn)
            continue
        entry    = index[tn]
        pdf_path = UPLOAD_DIR / entry.get("filename", "")
        page_num = entry.get("page", 1) - 1  # 0-indexed
        if not pdf_path.exists():
            missing.append(tn)
            continue
        try:
            reader = PdfReader(str(pdf_path))
            if 0 <= page_num < len(reader.pages):
                writer.add_page(reader.pages[page_num])
        except Exception:
            missing.append(tn)

    if not writer.pages:
        return jsonify({"error": "Nenhuma etiqueta encontrada no disco"}), 404

    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    today = date.today().strftime("%d.%m.%y")
    return send_file(buf, as_attachment=True,
                     download_name=f"pedidos_localizados_{today}.pdf",
                     mimetype="application/pdf")


def _split_pdf_columns(src_bytes, num_cols=None, with_selo=False):
    """Corta PDF com etiquetas lado a lado e retorna PDF com uma etiqueta por pÃ¡gina.
    Usa insert_pdf + manipulaÃ§Ã£o do content stream para preservar texto extraÃ­vel,
    idÃªntico ao comportamento do sistema corta-etiqueta original (pdf-lib).
    Constantes calibradas no sistema original: TOP_CROP=20 (PN), CONTENT_H=355 (TN)."""
    TARGET_W  = 283.46  # 100 mm
    TARGET_H  = 425.20  # 150 mm
    TOP_CROP  = 20
    CONTENT_H = 355
    AJUSTE_X  = -5      # AJUSTE_X_GLOBAL do sistema original

    # Carimbo "Conferido" â€” medidas do selo-config.ts (coords PDF: y=0 embaixo)
    SELO_X      = 165
    SELO_Y_PDF  = 205   # distÃ¢ncia da borda inferior
    SELO_W      = 80
    SELO_H      = 80

    selo_bytes = None
    if with_selo:
        selo_path = Path(__file__).parent / "conferido.png"
        if selo_path.exists():
            selo_bytes = selo_path.read_bytes()

    src_doc = fitz.open(stream=src_bytes, filetype="pdf")
    out_doc = fitz.open()

    for page_num in range(len(src_doc)):
        page = src_doc[page_num]
        w, h = page.rect.width, page.rect.height

        if num_cols is None:
            ratio = w / h if h else 1
            if   ratio > 2.5: cols = 4
            elif ratio > 1.2: cols = 4
            elif ratio > 0.6: cols = 2
            else:              cols = 1
        else:
            cols = int(num_cols)

        col_w = w / cols

        for col in range(cols):
            col_x0 = col * col_w
            col_x1 = (col + 1) * col_w

            if not page.get_text("blocks", clip=fitz.Rect(col_x0, 0, col_x1, h)):
                continue

            # Copia pÃ¡gina completa preservando fontes e recursos (texto extraÃ­vel)
            out_doc.insert_pdf(src_doc, from_page=page_num, to_page=page_num)
            out_page = out_doc[-1]

            # Matriz de transformaÃ§Ã£o: recorta coluna e escala para TARGET_W Ã— TARGET_H
            # Em coordenadas PDF nativas (y=0 embaixo):
            # clip inferior: h - TOP_CROP - CONTENT_H  (ex.: 612-20-355=237)
            scale_x = TARGET_W / col_w       # ~1.431 para 198â†’283.46
            scale_y = TARGET_H / CONTENT_H   # ~1.197 para 355â†’425.2
            pdf_y_bot = h - TOP_CROP - CONTENT_H
            start_x = col_x0 + AJUSTE_X      # AJUSTE_X_GLOBAL do sistema original
            tx = -start_x * scale_x
            ty = -pdf_y_bot * scale_y

            # Envolve o content stream original com a transformaÃ§Ã£o
            original = out_page.read_contents()
            new_content = (
                f"q {scale_x:.6f} 0 0 {scale_y:.6f} {tx:.6f} {ty:.6f} cm\n"
                .encode()
                + original
                + b"\nQ\n"
            )

            # Grava no primeiro xref de content e aponta /Contents para ele
            content_xrefs = out_page.get_contents()
            out_doc.update_stream(content_xrefs[0], new_content)
            out_page.set_contents(content_xrefs[0])

            # Define tamanho da pÃ¡gina de saÃ­da e remove CropBox residual
            out_page.set_mediabox(fitz.Rect(0, 0, TARGET_W, TARGET_H))
            out_doc.xref_set_key(out_page.xref, "CropBox", "null")

            # Overlay do carimbo "Conferido" (coords PDF y-up â†’ fitz y-down)
            if selo_bytes:
                selo_y0 = TARGET_H - SELO_Y_PDF - SELO_H  # topo em fitz
                out_page.insert_image(
                    fitz.Rect(SELO_X, selo_y0, SELO_X + SELO_W, selo_y0 + SELO_H),
                    stream=selo_bytes,
                    overlay=True,
                )

    try:
        cat_xref = out_doc.pdf_catalog()
        out_doc.xref_set_key(cat_xref, "ViewerPreferences", "<</PrintScaling /None>>")
    except Exception:
        pass

    result = out_doc.tobytes(garbage=4, deflate=True)
    out_doc.close()
    src_doc.close()
    return result


def _split_ml_full(src_bytes, labels_per_page=3, num_cols=4):
    """Corta PDF de etiquetas Full do Mercado Livre (grade cols Ã— linhas)
    e agrupa labels_per_page etiquetas por pÃ¡gina A4, empilhadas verticalmente."""
    OUT_W = 595.0   # A4 retrato
    OUT_H = 842.0

    src_doc = fitz.open(stream=src_bytes, filetype="pdf")
    cells = []  # (page_num, clip_rect)

    for page_num in range(len(src_doc)):
        page    = src_doc[page_num]
        w, h    = page.rect.width, page.rect.height
        col_w   = w / num_cols

        # Detecta linhas horizontais separadoras via desenhos do PDF
        y_lines = set()
        for d in page.get_drawings():
            r = d.get("rect")
            if r and r.width >= w * 0.7:
                y_lines.add(round(r.y0))
                y_lines.add(round(r.y1))

        y_bounds = sorted(y_lines)

        # Fallback: clusteriza por posiÃ§Ã£o Y dos blocos de texto
        if len(y_bounds) < 3:
            blocks = page.get_text("blocks")
            raw_ys = sorted(set(b[1] for b in blocks if b[4].strip()))
            y_bounds = [0.0]
            prev = -999.0
            for y in raw_ys:
                if y - prev > 15:
                    if y_bounds:
                        y_bounds.append((prev + y) / 2 if prev > 0 else y)
                    prev = y
            y_bounds.append(h)
            y_bounds = sorted(set(round(v) for v in y_bounds))

        for ri in range(len(y_bounds) - 1):
            y0, y1 = y_bounds[ri], y_bounds[ri + 1]
            if y1 - y0 < 10:
                continue
            for ci in range(num_cols):
                clip = fitz.Rect(ci * col_w, y0, (ci + 1) * col_w, y1)
                if page.get_text("text", clip=clip).strip():
                    cells.append((page_num, clip))

    out_doc  = fitz.open()
    slot_h   = OUT_H / labels_per_page

    for i in range(0, len(cells), labels_per_page):
        batch    = cells[i:i + labels_per_page]
        out_page = out_doc.new_page(width=OUT_W, height=OUT_H)
        for slot, (pn, clip) in enumerate(batch):
            dest = fitz.Rect(0, slot * slot_h, OUT_W, (slot + 1) * slot_h)
            out_page.show_pdf_page(dest, src_doc, pn, clip=clip, keep_proportion=False)

    try:
        cat = out_doc.pdf_catalog()
        out_doc.xref_set_key(cat, "ViewerPreferences", "<</PrintScaling /None>>")
    except Exception:
        pass

    result = out_doc.tobytes(garbage=4, deflate=True)
    out_doc.close()
    src_doc.close()
    return result


def _split_pdf_columns_temp(src_bytes):
    """Cortador TEMPORÁRIO para layout Dex Fit (landscape 842x595, 4 colunas
    agrupadas 2+2 com gap maior no meio). Recorta cada coluna e ESTICA para
    preencher uma etiqueta térmica 100x150mm (283x425), uma por página."""
    # X inicial de cada coluna + largura (calibrado p/ bater com OFICIAL DEX:
    # #=21, Produto=73, SKU=124, Var=176, Qtd=227 — alinhado aos limites do parser)
    COL_STARTS = [40, 232, 432, 624]
    COL_W      = 192          # largura do slot da etiqueta
    TARGET_W   = 283.46       # 100 mm
    TARGET_H   = 425.20       # 150 mm
    TOP_CROP   = 16           # corte do topo (fitz top-down)
    CONTENT_H  = 342          # altura real do conteúdo da etiqueta (y 16 → 358)

    src_doc = fitz.open(stream=src_bytes, filetype="pdf")
    out_doc = fitz.open()
    scale_y = TARGET_H / CONTENT_H

    for page_num in range(len(src_doc)):
        page = src_doc[page_num]
        h = page.rect.height
        for col_x0 in COL_STARTS:
            col_x1 = col_x0 + COL_W
            clip = fitz.Rect(col_x0, TOP_CROP, col_x1, TOP_CROP + CONTENT_H)
            if not page.get_text("blocks", clip=clip):
                continue
            scale_x = TARGET_W / COL_W
            pdf_y_bot = h - TOP_CROP - CONTENT_H   # borda inferior do recorte (coords PDF y-up)
            tx = -col_x0 * scale_x
            ty = -pdf_y_bot * scale_y

            out_doc.insert_pdf(src_doc, from_page=page_num, to_page=page_num)
            out_page = out_doc[-1]
            original = out_page.read_contents()
            new_content = (
                f"q {scale_x:.6f} 0 0 {scale_y:.6f} {tx:.6f} {ty:.6f} cm\n".encode()
                + original + b"\nQ\n"
            )
            content_xrefs = out_page.get_contents()
            out_doc.update_stream(content_xrefs[0], new_content)
            out_page.set_contents(content_xrefs[0])
            out_page.set_mediabox(fitz.Rect(0, 0, TARGET_W, TARGET_H))
            out_doc.xref_set_key(out_page.xref, "CropBox", "null")

            # Algumas etiquetas deste layout não trazem a coluna "#" do checklist
            # (necessária pro separador delimitar o item). Se faltar, injeta um "1"
            # na posição da coluna # alinhado à linha do produto.
            words = out_page.get_text("words")
            if not any(w[4].strip() == "#" for w in words):
                prod_hdr = next((w for w in words if w[4].strip() == "Produto"), None)
                if prod_hdr:
                    hy = prod_hdr[1]
                    data_ys = [w[1] for w in words if w[1] > hy + 1 and 55 < w[0] < 120]
                    if data_ys:
                        dy = min(data_ys)
                        out_page.insert_text(fitz.Point(20, dy + 6.5), "1",
                                             fontsize=7, color=(0, 0, 0))

    try:
        out_doc.xref_set_key(out_doc.pdf_catalog(), "ViewerPreferences", "<</PrintScaling /None>>")
    except Exception:
        pass
    result = out_doc.tobytes(garbage=4, deflate=True)
    out_doc.close()
    src_doc.close()
    return result


@app.route("/cortar/temp", methods=["POST"])
def cortar_temp():
    f = request.files.get("pdf")
    if not f or not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Envie um arquivo PDF válido"}), 400
    try:
        result = _split_pdf_columns_temp(f.read())
    except Exception as e:
        return jsonify({"error": f"Erro ao processar PDF: {e}"}), 500
    buf = io.BytesIO(result); buf.seek(0)
    safe_name = re.sub(r"[^\w.\-]", "_", f.filename.rsplit(".", 1)[0])
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"cortado_temp_{safe_name}.pdf")


@app.route("/cortar", methods=["POST"])
def cortar_etiquetas():
    f = request.files.get("pdf")
    if not f or not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Envie um arquivo PDF vÃ¡lido"}), 400
    cols_param = request.form.get("cols", "auto")
    num_cols   = None if cols_param == "auto" else int(cols_param)
    with_selo  = request.form.get("selo") == "1"
    try:
        result = _split_pdf_columns(f.read(), num_cols, with_selo=with_selo)
    except Exception as e:
        return jsonify({"error": f"Erro ao processar PDF: {e}"}), 500
    buf = io.BytesIO(result)
    buf.seek(0)
    safe_name = re.sub(r"[^\w.\-]", "_", f.filename.rsplit(".", 1)[0])
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True,
                     download_name=f"cortado_{safe_name}.pdf")


@app.route("/caixas/gerar", methods=["POST"])
def caixas_gerar():
    data     = request.get_json()
    date_iso = data.get("date", "")
    address  = data.get("address", "").strip()
    boxes    = data.get("boxes", [])
    total    = len(boxes)
    if not boxes:
        return jsonify({"error": "Nenhuma caixa informada"}), 400

    # Formata data de YYYY-MM-DD para DD/MM/YYYY
    try:
        from datetime import datetime as _dt
        date_display = _dt.strptime(date_iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        date_display = date_iso

    doc = fitz.open()
    W, H = 595, 842  # A4 retrato
    M    = 45        # margem horizontal

    for i, box in enumerate(boxes):
        num_pedidos = int(box.get("pedidos", 0))
        box_num     = i + 1
        page        = doc.new_page(width=W, height=H)

        # "N PEDIDOS" â€” negrito, muito grande
        r = fitz.Rect(M, 70, W - M, 230)
        page.insert_textbox(r, f"{num_pedidos} PEDIDOS",
                            fontsize=80, fontname="hebo",
                            align=fitz.TEXT_ALIGN_CENTER, color=(0, 0, 0))

        # "CAIXA X/Y" â€” negrito, muito grande
        r = fitz.Rect(M, 250, W - M, 430)
        page.insert_textbox(r, f"CAIXA {box_num}/{total}",
                            fontsize=80, fontname="hebo",
                            align=fitz.TEXT_ALIGN_CENTER, color=(0, 0, 0))

        # Data â€” itÃ¡lico, grande
        r = fitz.Rect(M, 445, W - M, 570)
        page.insert_textbox(r, date_display,
                            fontsize=54, fontname="heit",
                            align=fitz.TEXT_ALIGN_CENTER, color=(0, 0, 0))

        # EndereÃ§o â€” itÃ¡lico, menor, sublinhado manual
        if address:
            addr_size = 17
            r_addr = fitz.Rect(M, 640, W - M, 780)
            page.insert_textbox(r_addr, address,
                                fontsize=addr_size, fontname="heit",
                                align=fitz.TEXT_ALIGN_CENTER, color=(0, 0, 0))
            # Sublinhado: calcula largura do texto e desenha linha(s)
            lines = []
            words = address.split()
            current = ""
            line_w = W - 2 * M
            for word in words:
                test = (current + " " + word).strip()
                tw = fitz.get_text_length(test, fontname="heit", fontsize=addr_size)
                if tw <= line_w:
                    current = test
                else:
                    if current:
                        lines.append(current)
                    current = word
            if current:
                lines.append(current)
            line_h = addr_size * 1.35
            for li, line in enumerate(lines):
                tw  = fitz.get_text_length(line, fontname="heit", fontsize=addr_size)
                x0  = (W - tw) / 2
                x1  = x0 + tw
                y_u = 640 + (li + 1) * line_h + 1
                page.draw_line(fitz.Point(x0, y_u), fitz.Point(x1, y_u),
                               color=(0, 0, 0), width=0.7)

    buf = io.BytesIO()
    doc.save(buf)
    buf.seek(0)
    doc.close()
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name="caixas.pdf")


# â”€â”€ Estoque & Produtos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

ESTOQUE_FILE      = BASE_DIR / "estoque.json"
HIST_ESTOQUE_FILE = BASE_DIR / "historico_estoque.json"

_ESTOQUE_EXCLUDED = {"roupa", "roupas", "variacoes", "variaes"}

# Produtos padrÃ£o que compÃµem o estoque fÃ­sico
_ESTOQUE_DEFAULT = {
    "pura":  "Creatina Black Skull Pura",
    "turbo": "Creatina Black Skull Turbo",
    "dux":   "Creatina DUX",
}

def load_estoque():
    if ESTOQUE_FILE.exists():
        with open(ESTOQUE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {k: {"quantidade": 0, "display": v} for k, v in _ESTOQUE_DEFAULT.items()}

def save_estoque(e):
    with open(ESTOQUE_FILE, "w", encoding="utf-8") as f:
        json.dump(e, f, ensure_ascii=False, indent=2)

def load_hist_estoque():
    if HIST_ESTOQUE_FILE.exists():
        with open(HIST_ESTOQUE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_hist_estoque(h):
    with open(HIST_ESTOQUE_FILE, "w", encoding="utf-8") as f:
        json.dump(h, f, ensure_ascii=False, indent=2)

def _fmt_ts(ts):
    try:
        dt = datetime.fromisoformat(ts)
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return ts


@app.route("/estoque")
def estoque_page():
    estoque  = load_estoque()
    raw_hist = load_hist_estoque()
    hist = []
    for e in reversed(raw_hist[-200:]):
        e2 = dict(e)
        e2["ts_fmt"] = _fmt_ts(e.get("timestamp", ""))
        hist.append(e2)
    return render_template("estoque.html", estoque=estoque, historico=hist)


@app.route("/produtos")
def produtos_page():
    mappings      = load_mappings()
    brand_display = get_brand_display()
    grupos = {}
    for mk, info in mappings.items():
        cat = info.get("categoria", "?")
        sku, titulo_raw = (mk.split("|", 1) if "|" in mk else (mk, mk))
        titulo = info.get("titulo") or titulo_raw
        if cat not in grupos:
            grupos[cat] = {"display": brand_display.get(cat, cat.upper()), "itens": []}
        grupos[cat]["itens"].append({
            "key":      mk,
            "sku":      sku,
            "titulo":   titulo,
            "kit_size": info.get("kit_size", 1),
        })
    for g in grupos.values():
        g["itens"].sort(key=lambda x: x["titulo"].lower())
    ordered = []
    for cat in STANDARD_ORDER:
        if cat in grupos:
            ordered.append((cat, grupos.pop(cat)))
    for cat in sorted(grupos):
        ordered.append((cat, grupos[cat]))
    all_groups = [(cat, brand_display.get(cat, cat.upper()))
                  for cat, _ in ordered]
    # also include groups that exist only in brands but have no products yet
    for k, v in brand_display.items():
        if k not in [c for c, _ in all_groups] and k not in _ESTOQUE_EXCLUDED:
            all_groups.append((k, v))
    return render_template("produtos.html", grupos=ordered,
                           all_groups=all_groups, brand_display=brand_display)


@app.route("/api/estoque", methods=["GET"])
def api_estoque_get():
    return jsonify(load_estoque())


@app.route("/estoque-hist", methods=["GET"])
def estoque_hist():
    hist = load_hist_estoque()
    result = []
    for e in reversed(hist[-200:]):
        result.append(e)
    return jsonify(result)


@app.route("/api/produtos", methods=["GET"])
def api_produtos_get():
    mappings      = load_mappings()
    brand_display = get_brand_display()
    grupos = {}
    for mk, info in mappings.items():
        cat = info.get("categoria", "?")
        sku, titulo_raw = (mk.split("|", 1) if "|" in mk else (mk, mk))
        titulo = info.get("titulo") or titulo_raw
        if cat not in grupos:
            grupos[cat] = {"cat": cat, "display": brand_display.get(cat, cat.upper()), "itens": []}
        grupos[cat]["itens"].append({
            "key":      mk,
            "sku":      sku,
            "titulo":   titulo,
            "kit_size": info.get("kit_size", 1),
        })
    for g in grupos.values():
        g["itens"].sort(key=lambda x: x["titulo"].lower())
    # Incluir grupos vazios (criados mas sem produtos ainda)
    for cat, display in brand_display.items():
        if cat not in grupos and cat not in _ESTOQUE_EXCLUDED:
            grupos[cat] = {"cat": cat, "display": display, "itens": []}
    ordered = []
    for cat in STANDARD_ORDER:
        if cat in grupos:
            ordered.append(grupos.pop(cat))
    for cat in sorted(grupos):
        ordered.append(grupos[cat])
    return jsonify(ordered)


@app.route("/api/estoque/grupos", methods=["GET"])
def api_estoque_grupos():
    e = load_estoque()
    return jsonify([{"grupo": k, "display": v["display"], "quantidade": v["quantidade"]}
                    for k, v in e.items() if v.get("ativo", True)])


@app.route("/api/estoque/ativo", methods=["POST"])
def api_estoque_ativo():
    data   = request.json
    grupo  = data.get("grupo", "").strip()
    ativo  = bool(data.get("ativo", True))
    estoque = load_estoque()
    if grupo not in estoque:
        return jsonify({"error": "Grupo nÃ£o encontrado"}), 404
    estoque[grupo]["ativo"] = ativo
    save_estoque(estoque)
    status = "ativado" if ativo else "desativado"
    registrar_aud("estoque_ativo", f"Grupo '{grupo}' {status} no estoque")
    return jsonify({"ok": True})


@app.route("/api/todos-grupos", methods=["GET"])
def api_todos_grupos():
    brands  = load_brands()
    estoque = load_estoque()
    seen    = set()
    result  = []
    # Primeiro: grupos que jÃ¡ estÃ£o no estoque (pura/turbo/dux podem nÃ£o estar em brands.json)
    for slug, info in estoque.items():
        seen.add(slug)
        display = info.get("display") or brands.get(slug, slug)
        ativo   = info.get("ativo", True)
        result.append({"grupo": slug, "display": display, "ativo": ativo, "quantidade": info.get("quantidade", 0)})
    # Depois: marcas em brands.json que ainda nÃ£o estÃ£o no estoque
    for slug, display in brands.items():
        if slug not in seen:
            result.append({"grupo": slug, "display": display, "ativo": False, "quantidade": 0})
    return jsonify(result)


@app.route("/api/grupos/editar", methods=["POST"])
def api_grupos_editar():
    data        = request.json
    slug        = data.get("slug", "").strip()
    new_display = data.get("display", "").strip()
    if not slug or not new_display:
        return jsonify({"error": "Dados invÃ¡lidos"}), 400
    brands = load_brands()
    old_display = brands.get(slug, slug)
    brands[slug] = new_display
    save_brands(brands)
    estoque = load_estoque()
    if slug in estoque:
        old_display = estoque[slug].get("display", old_display)
        estoque[slug]["display"] = new_display
        save_estoque(estoque)
    registrar_aud("grupo_editar", f"Nome do grupo alterado: '{old_display}' â†’ '{new_display}'")
    return jsonify({"ok": True})


@app.route("/api/estoque/entrada", methods=["POST"])
def api_estoque_entrada():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data     = request.json
    itens    = data.get("itens", [])
    usuario  = session.get("user", "?")
    estoque  = load_estoque()
    hist     = load_hist_estoque()
    movimentos = []
    for item in itens:
        grupo = item.get("grupo", "")
        qtd   = int(item.get("quantidade", 0))
        if qtd <= 0:
            continue
        if grupo not in estoque:
            estoque[grupo] = {"quantidade": 0, "display": item.get("display", grupo)}
        estoque[grupo]["quantidade"] += qtd
        movimentos.append({"grupo": grupo, "display": estoque[grupo]["display"], "quantidade": qtd})
    if not movimentos:
        return jsonify({"error": "Nenhum item com quantidade vÃ¡lida"}), 400
    hist.append({
        "id":            secrets.token_hex(8),
        "timestamp":     datetime.now().isoformat(timespec="seconds"),
        "usuario":       usuario,
        "tipo":          "entrada",
        "itens":         movimentos,
        "justificativa": None,
        "contexto":      data.get("contexto", ""),
    })
    save_estoque(estoque)
    save_hist_estoque(hist)
    resumo = ", ".join(f"{m['display']} +{m['quantidade']}" for m in movimentos)
    registrar_aud("estoque_entrada", f"Entrada de estoque: {resumo}")
    return jsonify({"ok": True, "estoque": estoque})


@app.route("/api/estoque/saida-manual", methods=["POST"])
def api_estoque_saida_manual():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data          = request.json
    grupo         = (data.get("grupo") or "").strip()
    qtd           = int(data.get("quantidade", 0))
    justificativa = (data.get("justificativa") or "").strip()
    usuario       = session.get("user", "?")
    if not justificativa:
        return jsonify({"error": "Justificativa obrigatÃ³ria"}), 400
    if qtd <= 0:
        return jsonify({"error": "Quantidade invÃ¡lida"}), 400
    estoque = load_estoque()
    hist    = load_hist_estoque()
    if grupo not in estoque:
        return jsonify({"error": f"Produto nÃ£o encontrado: {grupo}"}), 400
    display = estoque[grupo]["display"]
    estoque[grupo]["quantidade"] -= qtd
    mov = {"grupo": grupo, "display": display, "quantidade": qtd}
    hist.append({
        "id":            secrets.token_hex(8),
        "timestamp":     datetime.now().isoformat(timespec="seconds"),
        "usuario":       usuario,
        "tipo":          "saida_manual",
        "itens":         [mov],
        "justificativa": justificativa,
        "contexto":      None,
    })
    save_estoque(estoque)
    save_hist_estoque(hist)
    registrar_aud("estoque_saida_manual", f"SaÃ­da manual: {display} -{qtd} â€” {justificativa}")
    return jsonify({"ok": True, "estoque": estoque})


@app.route("/api/estoque/saida-expedicao", methods=["POST"])
def api_estoque_saida_expedicao():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data     = request.json
    itens    = data.get("itens", [])
    contexto = data.get("contexto", "")
    usuario  = session.get("user", "?")
    estoque  = load_estoque()
    hist     = load_hist_estoque()
    movimentos = []
    for item in itens:
        grupo_real = item.get("grupo_real", "")
        qtd        = int(item.get("quantidade", 0))
        if qtd <= 0 or grupo_real not in estoque:
            continue
        estoque[grupo_real]["quantidade"] -= qtd
        movimentos.append({
            "grupo":            grupo_real,
            "display":          estoque[grupo_real]["display"],
            "quantidade":       qtd,
            "grupo_original":   item.get("grupo_original"),
            "display_original": item.get("display_original"),
        })
    if not movimentos:
        return jsonify({"error": "Nenhum item vÃ¡lido para descontar"}), 400
    hist.append({
        "id":            secrets.token_hex(8),
        "timestamp":     datetime.now().isoformat(timespec="seconds"),
        "usuario":       usuario,
        "tipo":          "saida_expedicao",
        "itens":         movimentos,
        "justificativa": None,
        "contexto":      contexto,
    })
    save_estoque(estoque)
    save_hist_estoque(hist)
    resumo = ", ".join(
        (f"{m['display']} -{m['quantidade']}" if m['grupo'] == m.get('grupo_original') else
         f"{m['display_original']} -{m['quantidade']} via {m['display']}")
        for m in movimentos
    )
    registrar_aud("estoque_expedicao", f"{contexto}: {resumo}")
    return jsonify({"ok": True, "estoque": estoque})


@app.route("/api/estoque/desfazer", methods=["POST"])
def api_estoque_desfazer():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data     = request.json
    entry_id = data.get("id", "")
    hist     = load_hist_estoque()
    entry    = next((e for e in hist if e.get("id") == entry_id), None)
    if not entry:
        return jsonify({"error": "Entrada nÃ£o encontrada"}), 404
    tipo    = entry.get("tipo", "")
    itens   = entry.get("itens", [])
    estoque = load_estoque()
    movimentos = []

    # SaÃ­da â†’ devolver ao estoque; Entrada/estorno â†’ remover do estoque
    if tipo in ("saida_manual", "saida_expedicao"):
        for item in itens:
            grupo = item.get("grupo", "")
            qtd   = int(item.get("quantidade", 0))
            if not grupo or qtd <= 0 or grupo not in estoque:
                continue
            estoque[grupo]["quantidade"] += qtd
            movimentos.append({"grupo": grupo, "display": estoque[grupo].get("display", grupo), "quantidade": qtd, "sinal": "+"})
        novo_tipo = "estorno"
    elif tipo in ("entrada", "estorno_historico"):
        for item in itens:
            grupo = item.get("grupo", "")
            qtd   = int(item.get("quantidade", 0))
            if not grupo or qtd <= 0 or grupo not in estoque:
                continue
            estoque[grupo]["quantidade"] -= qtd
            movimentos.append({"grupo": grupo, "display": estoque[grupo].get("display", grupo), "quantidade": qtd, "sinal": "-"})
        novo_tipo = "estorno_entrada"
    else:
        return jsonify({"error": f"Tipo '{tipo}' nÃ£o suporta desfazer"}), 400

    if not movimentos:
        return jsonify({"error": "Nenhuma movimentaÃ§Ã£o vÃ¡lida para desfazer"}), 400

    justificativa = f"Desfazer: {entry.get('justificativa') or entry.get('contexto') or tipo}"
    hist.append({
        "id":            secrets.token_hex(8),
        "timestamp":     datetime.now().isoformat(timespec="seconds"),
        "usuario":       session.get("user", "?"),
        "tipo":          novo_tipo,
        "itens":         [{"grupo": m["grupo"], "display": m["display"], "quantidade": m["quantidade"]} for m in movimentos],
        "justificativa": justificativa,
        "contexto":      entry_id,
    })
    save_estoque(estoque)
    save_hist_estoque(hist)
    resumo = ", ".join(f"{m['display']} {m['sinal']}{m['quantidade']}" for m in movimentos)
    registrar_aud("estoque_desfazer", f"Desfazer movimentaÃ§Ã£o: {resumo}")
    return jsonify({"ok": True, "estoque": estoque})


@app.route("/api/produtos/mover", methods=["POST"])
def api_produtos_mover():
    data  = request.json
    key   = data.get("key")
    grupo = data.get("grupo")
    if not key or not grupo:
        return jsonify({"error": "Dados invÃ¡lidos"}), 400
    mappings = load_mappings()
    if key not in mappings:
        return jsonify({"error": "Produto nÃ£o encontrado"}), 404
    titulo = mappings[key].get("titulo", key)
    origem = mappings[key].get("categoria", "?")
    mappings[key]["categoria"] = grupo
    save_mappings(mappings)
    registrar_aud("prod_mover", f"Produto '{titulo}' movido de '{origem}' para '{grupo}'")
    return jsonify({"ok": True})


@app.route("/api/produtos/novo", methods=["POST"])
def api_produtos_novo():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data     = request.json
    sku      = (data.get("sku") or "").strip()
    titulo   = (data.get("titulo") or "").strip()
    grupo    = (data.get("grupo") or "").strip()
    kit_size = int(data.get("kit_size", 1))
    if not sku or not titulo or not grupo:
        return jsonify({"error": "SKU, tÃ­tulo e grupo sÃ£o obrigatÃ³rios"}), 400
    mappings = load_mappings()
    key      = mapping_key(titulo, sku)
    if key in mappings:
        return jsonify({"error": "Produto jÃ¡ cadastrado com este SKU+tÃ­tulo"}), 400
    mappings[key] = {"categoria": grupo, "kit_size": kit_size, "titulo": titulo}
    save_mappings(mappings)
    registrar_aud("prod_novo", f"Novo produto: {titulo} ({sku}) â†’ {grupo}")
    return jsonify({"ok": True, "key": key})


@app.route("/api/produtos/deletar", methods=["POST"])
def api_produtos_deletar():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data = request.json
    key  = data.get("key")
    if not key:
        return jsonify({"error": "Key obrigatÃ³ria"}), 400
    mappings = load_mappings()
    if key not in mappings:
        return jsonify({"error": "Produto nÃ£o encontrado"}), 404
    titulo = mappings[key].get("titulo", key)
    del mappings[key]
    save_mappings(mappings)
    registrar_aud("prod_deletar", f"Produto removido: {titulo}")
    return jsonify({"ok": True})


@app.route("/api/produtos/editar", methods=["POST"])
def api_produtos_editar():
    data      = request.json
    key       = data.get("key")
    new_titulo = (data.get("titulo") or "").strip()
    if not key or not new_titulo:
        return jsonify({"error": "Dados invÃ¡lidos"}), 400
    mappings = load_mappings()
    if key not in mappings:
        return jsonify({"error": "Produto nÃ£o encontrado"}), 404
    entry    = mappings.pop(key)
    old_titulo = entry.get("titulo", "")
    entry["titulo"] = new_titulo
    # Regenerate key with new title (keep same SKU part from key)
    sku = key.split("|||")[1] if "|||" in key else key
    new_key = mapping_key(new_titulo, sku)
    mappings[new_key] = entry
    save_mappings(mappings)
    registrar_aud("prod_editar", f"Produto renomeado: '{old_titulo}' â†’ '{new_titulo}'")
    return jsonify({"ok": True, "new_key": new_key})


@app.route("/api/grupos/novo", methods=["POST"])
def api_grupos_novo():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data    = request.json
    slug    = (data.get("slug") or "").strip().lower()
    display = (data.get("display") or "").strip()
    if not slug or not display:
        return jsonify({"error": "Slug e nome sÃ£o obrigatÃ³rios"}), 400
    if not re.match(r'^[a-z0-9]+$', slug):
        return jsonify({"error": "Slug: apenas letras minÃºsculas e nÃºmeros"}), 400
    brands  = load_brands()
    estoque = load_estoque()
    brands[slug] = display
    save_brands(brands)
    if slug not in estoque:
        estoque[slug] = {"quantidade": 0, "display": display}
        save_estoque(estoque)
    registrar_aud("grupo_novo", f"Novo grupo criado: '{display}' (slug: {slug})")
    return jsonify({"ok": True})


@app.route("/api/grupos/excluir", methods=["POST"])
def api_grupos_excluir():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    slug = (request.json.get("slug") or "").strip().lower()
    if not slug:
        return jsonify({"error": "Slug Ã© obrigatÃ³rio"}), 400
    mappings = load_mappings()
    if any(v.get("categoria") == slug for v in mappings.values()):
        return jsonify({"error": "Grupo tem produtos. Mova-os antes de excluir."}), 400
    estoque = load_estoque()
    if estoque.get(slug, {}).get("quantidade", 0) > 0:
        return jsonify({"error": "Grupo tem estoque. Zere antes de excluir."}), 400
    brands = load_brands()
    display = brands.pop(slug, slug)
    save_brands(brands)
    estoque.pop(slug, None)
    save_estoque(estoque)
    registrar_aud("grupo_excluir", f"Grupo excluÃ­do: '{display}' (slug: {slug})")
    return jsonify({"ok": True})


# â”€â”€ API TESTE â€” Mercado Livre Integration â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

import urllib.request as _urllib_req
import urllib.parse   as _urllib_parse
import urllib.error   as _urllib_err

_ML_CLIENT_ID     = "6915156252689047"
_ML_CLIENT_SECRET = "9lN5228liAvl7mrY8IdGe1Jp2jJrqvsp"
_ML_TOKEN_FILE    = BASE_DIR / "ml_tokens.json"
_ML_REDIRECT_URI  = "https://www.sistemaomni.com.br/api-teste/ml/callback"
_ML_AUTH_URL      = "https://auth.mercadolivre.com.br/authorization"
_ML_TOKEN_URL     = "https://api.mercadolibre.com/oauth/token"
_ML_API_BASE      = "https://api.mercadolibre.com"
_ML_STATE_FILE    = BASE_DIR / "ml_oauth_state.json"


def _ml_load_tokens():
    if _ML_TOKEN_FILE.exists():
        return json.loads(_ML_TOKEN_FILE.read_text(encoding="utf-8"))
    return {}


def _ml_save_tokens(tokens):
    _ML_TOKEN_FILE.write_text(json.dumps(tokens, ensure_ascii=False, indent=2), encoding="utf-8")


def _ml_exchange_token(grant_type, **kwargs):
    payload = _urllib_parse.urlencode({
        "grant_type":    grant_type,
        "client_id":     _ML_CLIENT_ID,
        "client_secret": _ML_CLIENT_SECRET,
        **kwargs,
    }).encode()
    req = _urllib_req.Request(
        _ML_TOKEN_URL, data=payload, method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    )
    try:
        with _urllib_req.urlopen(req, timeout=10) as r:
            return json.loads(r.read()), None
    except _urllib_err.HTTPError as e:
        return None, f"HTTP {e.code}: {e.read().decode()}"
    except Exception as e:
        return None, str(e)


def _ml_api(seller_id, path, params=None, retry=True):
    tokens = _ml_load_tokens()
    t = tokens.get(str(seller_id), {})
    access_token = t.get("access_token")
    if not access_token:
        return None, "Conta nÃ£o conectada"
    url = _ML_API_BASE + path
    if params:
        url += "?" + _urllib_parse.urlencode(params)
    req = _urllib_req.Request(url, headers={"Authorization": f"Bearer {access_token}"})
    try:
        with _urllib_req.urlopen(req, timeout=15) as r:
            return json.loads(r.read()), None
    except _urllib_err.HTTPError as e:
        if e.code == 401 and retry:
            refresh = t.get("refresh_token")
            if refresh:
                new_data, _ = _ml_exchange_token("refresh_token", refresh_token=refresh)
                if new_data:
                    tokens[str(seller_id)].update(new_data)
                    _ml_save_tokens(tokens)
                    return _ml_api(seller_id, path, params, retry=False)
        try:    body = e.read().decode()
        except: body = ""
        return None, f"HTTP {e.code}: {body}"
    except Exception as e:
        return None, str(e)


@app.route("/api-teste/ml/auth")
def api_teste_ml_auth():
    state = secrets.token_hex(10)
    _ML_STATE_FILE.write_text(json.dumps({"state": state, "ts": _time.time()}), encoding="utf-8")
    params = _urllib_parse.urlencode({
        "response_type": "code",
        "client_id":     _ML_CLIENT_ID,
        "redirect_uri":  _ML_REDIRECT_URI,
        "state":         state,
    })
    return redirect(f"{_ML_AUTH_URL}?{params}")


@app.route("/api-teste/ml/callback")
def api_teste_ml_callback():
    code  = request.args.get("code")
    error = request.args.get("error")
    _ML_STATE_FILE.unlink(missing_ok=True)

    if error or not code:
        (BASE_DIR / "ml_debug.log").write_text(f"callback error: error={error} code={code}", encoding="utf-8")
        return redirect("/?ml_err=1")

    token_data, err = _ml_exchange_token(
        "authorization_code", code=code, redirect_uri=_ML_REDIRECT_URI
    )
    if not token_data:
        (BASE_DIR / "ml_debug.log").write_text(f"token exchange failed: {err}", encoding="utf-8")
        return redirect("/?ml_err=1")

    seller_id = str(token_data.get("user_id", ""))
    if not seller_id:
        (BASE_DIR / "ml_debug.log").write_text(f"no user_id in token: {token_data}", encoding="utf-8")
        return redirect("/?ml_err=1")

    tokens = _ml_load_tokens()
    tokens[seller_id] = token_data
    user_data, _ = _ml_api(seller_id, f"/users/{seller_id}")
    if user_data:
        tokens[seller_id]["nickname"] = user_data.get("nickname", "")
        tokens[seller_id]["email"]    = user_data.get("email", "")
    _ml_save_tokens(tokens)
    (BASE_DIR / "ml_debug.log").write_text(f"ok: seller_id={seller_id}", encoding="utf-8")
    return redirect("/?ml_ok=1")


@app.route("/api-teste/ml/accounts")
def api_teste_ml_accounts():
    tokens = _ml_load_tokens()
    accounts = [{"seller_id": sid,
                 "nickname":  t.get("nickname", sid),
                 "email":     t.get("email", "")}
                for sid, t in tokens.items()]
    return jsonify({"accounts": accounts})


@app.route("/api-teste/ml/disconnect", methods=["POST"])
def api_teste_ml_disconnect():
    sid = str(request.json.get("seller_id", ""))
    tokens = _ml_load_tokens()
    tokens.pop(sid, None)
    _ml_save_tokens(tokens)
    return jsonify({"ok": True})


@app.route("/api-teste/ml/data/user")
def api_teste_ml_data_user():
    sid = request.args.get("seller_id")
    d, e = _ml_api(sid, f"/users/{sid}")
    if e: return jsonify({"error": e}), 400
    return jsonify(d)


@app.route("/api-teste/ml/data/orders")
def api_teste_ml_data_orders():
    sid = request.args.get("seller_id")
    d, e = _ml_api(sid, "/orders/search", {"seller": sid, "sort": "date_desc", "limit": 20})
    if e: return jsonify({"error": e}), 400
    return jsonify(d)


@app.route("/api-teste/ml/data/items")
def api_teste_ml_data_items():
    sid = request.args.get("seller_id")
    d, e = _ml_api(sid, f"/users/{sid}/items/search", {"status": "active", "limit": 20})
    if e: return jsonify({"error": e}), 400
    item_ids = d.get("results", [])
    if item_ids:
        ids_str = ",".join(item_ids[:20])
        details, _ = _ml_api(sid, "/items", {
            "ids": ids_str,
            "attributes": "id,title,price,available_quantity,sold_quantity,status,thumbnail"
        })
        if details and isinstance(details, list):
            d["item_details"] = [x.get("body", x) for x in details]
    return jsonify(d)


@app.route("/api-teste/ml/data/billing")
def api_teste_ml_data_billing():
    sid = request.args.get("seller_id")
    # Pega Ãºltimos 50 pedidos e agrupa por mÃªs
    d, e = _ml_api(sid, "/orders/search", {"seller": sid, "sort": "date_desc", "limit": 50})
    if e: return jsonify({"error": e}), 400
    orders = d.get("results", [])
    months = {}
    for o in orders:
        dt = o.get("date_created", "")[:7]  # YYYY-MM
        if not dt: continue
        if dt not in months:
            months[dt] = {"mes": dt, "pedidos": 0, "total": 0.0, "pago": 0.0}
        months[dt]["pedidos"] += 1
        months[dt]["total"] += o.get("total_amount", 0) or 0
        months[dt]["pago"]  += o.get("paid_amount", 0) or 0
    return jsonify({"meses": sorted(months.values(), key=lambda x: x["mes"], reverse=True),
                    "total_orders": d.get("paging", {}).get("total", 0)})


@app.route("/api-teste/ml/data/ads")
def api_teste_ml_data_ads():
    sid = request.args.get("seller_id")
    d, e = _ml_api(sid, f"/advertising/product_ads/advertisers/{sid}")
    if e:
        d2, e2 = _ml_api(sid, "/advertising/product_ads/advertisers", {"user_id": sid})
        if e2: return jsonify({"error": e}), 400
        return jsonify(d2)
    return jsonify(d)


@app.route("/api-teste/ml/data/reputation")
def api_teste_ml_data_reputation():
    sid = request.args.get("seller_id")
    d, e = _ml_api(sid, f"/users/{sid}")
    if e: return jsonify({"error": e}), 400
    return jsonify({"seller_reputation": d.get("seller_reputation", {}),
                    "status": d.get("status", {}),
                    "points": d.get("points", 0),
                    "raw": d.get("seller_reputation", {})})


@app.route("/api-teste/ml/data/payments")
def api_teste_ml_data_payments():
    sid = request.args.get("seller_id")
    d, e = _ml_api(sid, "/orders/search", {
        "seller": sid, "sort": "date_desc", "limit": 20,
        "payment.status": "approved"
    })
    if e: return jsonify({"error": e}), 400
    return jsonify(d)


# â”€â”€ TikTok Shop â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_TT_HISTORY_FILE  = BASE_DIR / "tiktok_history.json"
_TT_MAPPINGS_FILE = BASE_DIR / "tiktok_mappings.json"
_TT_SIZES = ["GGG", "EGG", "GG", "EG", "XG", "PP", "G", "M", "P"]
_TT_PRODUCT_LABELS = {
    "top_calca":     "Top + CalÃ§a",
    "top_short":     "Top + Short",
    "bermuda":       "Bermuda Masculina",
    "kit_camisetas": "Kit Camisetas",
}


def _tt_load_history():
    if _TT_HISTORY_FILE.exists():
        return json.loads(_TT_HISTORY_FILE.read_text(encoding="utf-8"))
    return []

def _tt_save_history(h):
    _TT_HISTORY_FILE.write_text(json.dumps(h, ensure_ascii=False, indent=2), encoding="utf-8")

def _tt_load_mappings():
    if _TT_MAPPINGS_FILE.exists():
        return json.loads(_TT_MAPPINGS_FILE.read_text(encoding="utf-8"))
    return {}

def _tt_save_mappings(m):
    _TT_MAPPINGS_FILE.write_text(json.dumps(m, ensure_ascii=False, indent=2), encoding="utf-8")

def _tt_normalize(title):
    return re.sub(r'\s+', ' ', title.lower().strip())

def _tt_auto_detect(name):
    t = name.lower()
    if "bermuda" in t:
        return "bermuda"
    if "camiseta" in t:
        return "kit_camisetas"
    if "short" in t:
        return "top_short"
    if "calÃ§a" in t or "calca" in t:
        return "top_calca"
    return None

def _tt_parse_label_page(page):
    """Extrai a data de envio da pÃ¡gina de etiqueta TikTok.
    O formato no PDF Ã©:  Data de envio:\n2026-04-24 15:49
    A data fica na linha SEGUINTE ao label."""
    lines = [l.strip() for l in page.get_text().split('\n')]
    for i, line in enumerate(lines):
        if "Data de envio" in line:
            # Tenta a mesma linha (depois do ":")
            after = line.split(":")[-1].strip() if ":" in line else ""
            m = re.search(r'(\d{4}-\d{2}-\d{2})', after)
            if m:
                return m.group(1)
            # A data estÃ¡ na prÃ³xima linha
            for j in range(i + 1, min(i + 4, len(lines))):
                m = re.search(r'(\d{4}-\d{2}-\d{2})', lines[j])
                if m:
                    return m.group(1)
    return ""

def _tt_parse_packing_list(page):
    wlist = [(round(x0,1), round(y0,1), w) for x0,y0,x1,y1,w,*_ in page.get_text("words")]

    # Locate header row: "Product Name SKU Seller SKU Qty"
    header_y = None
    col_xs   = {}
    for i, (x0, y0, w) in enumerate(wlist):
        if w == "Product":
            for j in range(i+1, min(i+6, len(wlist))):
                x2, y2, w2 = wlist[j]
                if w2 == "Name" and abs(y2 - y0) < 3:
                    header_y = y0
                    col_xs["produto"] = x0
                    break
        if header_y:
            break
    if header_y is None:
        return None

    sku_count = 0
    for x0, y0, w in wlist:
        if abs(y0 - header_y) > 3:
            continue
        if w == "SKU":
            sku_count += 1
            if sku_count == 1:
                col_xs["sku"] = x0
            elif "sku_code" not in col_xs:
                col_xs["sku_code"] = x0
        elif w == "Seller" and "sku_code" not in col_xs:
            col_xs["sku_code"] = x0
        elif w == "Qty":
            col_xs["qty"] = x0

    for k in ["produto", "sku", "qty"]:
        if k not in col_xs:
            return None
    if "sku_code" not in col_xs:
        col_xs["sku_code"] = col_xs["qty"] - 60

    col_order = sorted((v, k) for k, v in col_xs.items()
                       if k in ["produto", "sku", "sku_code", "qty"])

    def col_of(x):
        r = col_order[0][1]
        for cx, cn in col_order:
            if x >= cx - 8:
                r = cn
        return r

    # Metadata
    meta = {"tracking": "", "order_id": "", "date": ""}
    for line in page.get_text().split('\n'):
        line = line.strip()
        if line.startswith("Tracking number:"):
            meta["tracking"] = line.split(":", 1)[1].strip()
        elif line.startswith("Order ID:") and not meta["order_id"]:
            meta["order_id"] = line.split(":", 1)[1].strip()
        elif line.startswith("Created Time:"):
            meta["date"] = line.split(":", 1)[1].strip()

    data = [{"x": x0, "y": y0, "word": w, "col": col_of(x0)}
            for x0, y0, w in wlist if y0 > header_y + 2]

    y_groups = {}
    for d in data:
        yk = round(d["y"] / 2) * 2
        y_groups.setdefault(yk, []).append(d)

    items, current = [], None
    for yk in sorted(y_groups):
        row = y_groups[yk]
        by_col = {}
        for d in row:
            by_col.setdefault(d["col"], []).append(d["word"])
        if any(d["word"] in ["Total:", "TikTok"] for d in row):
            if current:
                items.append(current)
                current = None
            break
        if "qty" in by_col:
            if current:
                items.append(current)
            qty_raw = by_col["qty"][0]
            current = {
                "produto":    " ".join(by_col.get("produto", [])),
                "sku_raw":    " ".join(by_col.get("sku", [])),
                "seller_sku": " ".join(by_col.get("sku_code", [])),
                "qty":        int(qty_raw) if qty_raw.isdigit() else 1,
            }
        elif current:
            if "produto" in by_col:
                current["produto"] += " " + " ".join(by_col["produto"])
            if "sku" in by_col:
                extra = " ".join(by_col["sku"])
                current["sku_raw"] = (current["sku_raw"] + " " + extra).strip() if current["sku_raw"] else extra
            if "sku_code" in by_col and not current["seller_sku"]:
                current["seller_sku"] = " ".join(by_col["sku_code"])
    if current:
        items.append(current)

    for item in items:
        sku = item.pop("sku_raw", "").strip()
        seller = item.get("seller_sku", "").strip()
        if "," in sku:
            cp, sp = sku.rsplit(",", 1)
            item["color"] = cp.strip()
            raw_size = sp.strip().upper()
            item["size"] = raw_size.split()[0] if raw_size else ""
        else:
            item["color"] = sku
            item["size"]  = ""
        # Remove seller_sku contamination from color/size
        if seller:
            for field in ("color", "size"):
                val = item.get(field, "")
                if val.upper().endswith(seller.upper()):
                    item[field] = val[:-len(seller)].strip(" -").strip()
                elif seller.upper() in val.upper():
                    item[field] = re.sub(re.escape(seller), "", val, flags=re.IGNORECASE).strip(" -").strip()

    meta["items"] = items
    return meta


def _tt_parse_pdfs(filenames):
    orders = []
    for fname in filenames:
        path = UPLOAD_DIR / fname
        if not path.exists():
            continue
        doc = fitz.open(str(path))
        for i in range(0, len(doc) - 1, 2):
            ship_date  = _tt_parse_label_page(doc[i])
            order_data = _tt_parse_packing_list(doc[i + 1])
            if not order_data:
                continue
            for item in order_data.get("items", []):
                orders.append({
                    "tracking":   order_data.get("tracking", ""),
                    "order_id":   order_data.get("order_id", ""),
                    "created":    order_data.get("date", ""),
                    "ship_date":  ship_date,
                    "produto":    item.get("produto", "").strip(),
                    "color":      item.get("color", ""),
                    "size":       item.get("size", ""),
                    "seller_sku": item.get("seller_sku", ""),
                    "qty":        item.get("qty", 1),
                })
        doc.close()
    return orders


def _tt_get_unknowns(orders, mappings):
    seen, unknowns = set(), []
    for o in orders:
        name = o.get("produto", "").strip()
        norm = _tt_normalize(name)
        if not name or norm in seen:
            continue
        seen.add(norm)
        if norm not in mappings and not _tt_auto_detect(name):
            unknowns.append({"key": norm, "label": name})
    return unknowns


def _tt_apply_mappings(orders, mappings):
    for o in orders:
        norm = _tt_normalize(o.get("produto", ""))
        o["product_type"] = mappings.get(norm) or _tt_auto_detect(o.get("produto", "")) or "top_calca"
    return orders


@app.route("/tiktok/upload", methods=["POST"])
def tiktok_upload():
    files = request.files.getlist("pdfs")
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    saved = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        safe  = re.sub(r'[^\w.\-]', '_', f.filename)
        fname = "tiktok_" + safe
        f.save(str(UPLOAD_DIR / fname))
        saved.append(fname)
    if not saved:
        return jsonify({"error": "Nenhum PDF vÃ¡lido"}), 400
    orders   = _tt_parse_pdfs(saved)
    mappings = _tt_load_mappings()
    unknowns = _tt_get_unknowns(orders, mappings)
    return jsonify({
        "filenames":    saved,
        "total_orders": len(set(o["order_id"] for o in orders if o["order_id"])),
        "total_items":  sum(o["qty"] for o in orders),
        "unknowns":     unknowns,
    })


@app.route("/tiktok/save-mapping", methods=["POST"])
def tiktok_save_mapping():
    mappings = _tt_load_mappings()
    mappings.update(request.json.get("mappings", {}))
    _tt_save_mappings(mappings)
    return jsonify({"ok": True})


@app.route("/tiktok/process", methods=["POST"])
def tiktok_process():
    data      = request.json
    filenames = data.get("filenames", [])
    if not filenames:
        return jsonify({"error": "Nenhum arquivo"}), 400
    orders   = _tt_parse_pdfs(filenames)
    mappings = _tt_load_mappings()
    unknowns = _tt_get_unknowns(orders, mappings)
    if unknowns:
        return jsonify({"error": "Produtos nÃ£o classificados", "unknowns": unknowns}), 400
    orders = _tt_apply_mappings(orders, mappings)
    by_product, by_color, by_size, by_ship_date = {}, {}, {}, {}
    ship_dates = []
    for o in orders:
        qty = o.get("qty", 1)
        pt  = o.get("product_type", "top_calca")
        c   = (o.get("color") or "").upper()
        s   = (o.get("size")  or "").upper()
        sd  = (o.get("ship_date") or "")[:10]   # YYYY-MM-DD
        by_product[pt]            = by_product.get(pt, 0) + qty
        if c: by_color[c]         = by_color.get(c, 0) + qty
        if s: by_size[s]          = by_size.get(s, 0) + qty
        if sd:
            by_ship_date[sd]      = by_ship_date.get(sd, 0) + 1
            ship_dates.append(sd)
    ship_dates_sorted = sorted(set(ship_dates))
    date_range = {"from": ship_dates_sorted[0], "to": ship_dates_sorted[-1]} if ship_dates_sorted else {}
    return jsonify({
        "orders":        orders,
        "total_orders":  len(set(o["order_id"] for o in orders if o["order_id"])),
        "total_items":   sum(o["qty"] for o in orders),
        "by_product":    by_product,
        "by_color":      by_color,
        "by_size":       by_size,
        "by_ship_date":  dict(sorted(by_ship_date.items())),
        "date_range":    date_range,
    })


@app.route("/tiktok/lacrar", methods=["POST"])
def tiktok_lacrar():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    data       = request.json
    history    = _tt_load_history()
    entry_date = data.get("date") or date.today().isoformat()
    entry = {
        "id":           datetime.now().strftime("%Y%m%d-%H%M%S"),
        "date":         entry_date,
        "timestamp":    datetime.now().isoformat(),
        "total_orders": data.get("total_orders", 0),
        "total_items":  data.get("total_items", 0),
        "by_product":   data.get("by_product", {}),
        "by_color":     data.get("by_color", {}),
        "by_size":      data.get("by_size", {}),
        "orders":       data.get("orders", []),
    }
    history.append(entry)
    _tt_save_history(history)
    for fname in data.get("filenames", []):
        p = UPLOAD_DIR / fname
        if p.exists():
            try: p.unlink()
            except: pass
    return jsonify({"ok": True, "id": entry["id"]})


@app.route("/tiktok/dashboard-data")
def tiktok_dashboard_data():
    history   = _tt_load_history()
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    filter_pt = request.args.get("product_type", "")
    filter_c  = request.args.get("color", "").upper()
    filter_s  = request.args.get("size", "").upper()
    total_orders, total_items = 0, 0
    by_product, by_color, by_size, by_date, by_ship_date = {}, {}, {}, {}, {}
    all_colors, all_sizes = set(), set()
    for entry in history:
        d = entry.get("date", "")
        if date_from and d < date_from: continue
        if date_to   and d > date_to:   continue
        for o in entry.get("orders", []):
            c  = (o.get("color") or "").upper()
            s  = (o.get("size")  or "").upper()
            sd = (o.get("ship_date") or "")[:10]
            if c: all_colors.add(c)
            if s: all_sizes.add(s)
        seen_orders = set()
        for o in entry.get("orders", []):
            if filter_pt and o.get("product_type") != filter_pt: continue
            if filter_c  and (o.get("color") or "").upper() != filter_c: continue
            if filter_s  and (o.get("size")  or "").upper() != filter_s: continue
            qty = o.get("qty", 1)
            oid = o.get("order_id") or o.get("tracking", "")
            total_items  += qty
            if oid and oid not in seen_orders:
                seen_orders.add(oid)
                total_orders += 1
            elif not oid:
                total_orders += 1
            pt  = o.get("product_type", "")
            c   = (o.get("color") or "").upper()
            s   = (o.get("size")  or "").upper()
            sd  = (o.get("ship_date") or "")[:10]
            by_product[pt]         = by_product.get(pt, 0) + qty
            if c: by_color[c]      = by_color.get(c, 0) + qty
            if s: by_size[s]       = by_size.get(s, 0) + qty
            if sd: by_date[sd]     = by_date.get(sd, 0) + 1
            if sd: by_ship_date[sd]= by_ship_date.get(sd, 0) + 1
    return jsonify({
        "total_orders":  total_orders,
        "total_items":   total_items,
        "by_product":    by_product,
        "by_color":      dict(sorted(by_color.items(), key=lambda x: -x[1])),
        "by_size":       dict(sorted(by_size.items())),
        "by_date":       dict(sorted(by_date.items(), reverse=True)),
        "by_ship_date":  dict(sorted(by_ship_date.items())),
        "all_colors":    sorted(all_colors),
        "all_sizes":     sorted(all_sizes),
    })


@app.route("/tiktok/history")
def tiktok_history_list():
    history = _tt_load_history()
    return jsonify({"entries": list(reversed(history))})


@app.route("/tiktok/history/delete", methods=["POST"])
def tiktok_history_delete():
    if not _valida_pin_session():
        return jsonify({"error": "ConfirmaÃ§Ã£o de PIN necessÃ¡ria"}), 403
    eid     = request.json.get("id")
    history = _tt_load_history()
    history = [e for e in history if e.get("id") != eid]
    _tt_save_history(history)
    return jsonify({"ok": True})


@app.route("/tiktok/debug-pdf", methods=["POST"])
def tiktok_debug_pdf():
    """Endpoint temporÃ¡rio para diagnosticar extraÃ§Ã£o de texto do PDF."""
    f = request.files.get("pdf")
    if not f:
        return jsonify({"error": "Nenhum arquivo"}), 400
    import tempfile, os
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name
    doc = fitz.open(tmp_path)
    result = []
    for i in range(min(4, len(doc))):
        page = doc[i]
        text = page.get_text()
        words_sample = page.get_text("words")[:30]
        ship = _tt_parse_label_page(page)
        result.append({
            "page": i,
            "text_first_500": text[:500],
            "words_sample": [w[4] for w in words_sample],
            "ship_date_extracted": ship,
        })
    doc.close()
    os.unlink(tmp_path)
    return jsonify(result)


# â”€â”€ Shopee API Integration â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
import hmac as _hmac

_SHP_PARTNER_ID  = 1234628
_SHP_PARTNER_KEY = "776267475a4d7663756c624b6a45584f4d64784c594a624d6f6d78416e6e"
_SHP_TOKEN_FILE  = BASE_DIR / "shopee_tokens.json"
_SHP_REDIRECT    = "https://www.sistemaomni.com.br/api-teste/shopee/callback"
_SHP_BASE        = "https://partner.test-stable.shopeemobile.com"


def _shp_sign(path, timestamp, access_token="", shop_id=""):
    base = f"{_SHP_PARTNER_ID}{path}{timestamp}{access_token}{shop_id}"
    key = bytes.fromhex(_SHP_PARTNER_KEY)
    return _hmac.new(key, base.encode(), hashlib.sha256).hexdigest()


def _shp_load_tokens():
    if _SHP_TOKEN_FILE.exists():
        return json.loads(_SHP_TOKEN_FILE.read_text(encoding="utf-8"))
    return {}


def _shp_save_tokens(tokens):
    _SHP_TOKEN_FILE.write_text(json.dumps(tokens, ensure_ascii=False, indent=2), encoding="utf-8")


def _shp_refresh_token(shop_id, refresh_token):
    path = "/api/v2/auth/access_token/get"
    ts = int(_time.time())
    sign = _shp_sign(path, ts)
    url = _SHP_BASE + path + "?" + _urllib_parse.urlencode({
        "partner_id": _SHP_PARTNER_ID, "timestamp": ts, "sign": sign,
    })
    payload = json.dumps({"refresh_token": refresh_token, "shop_id": shop_id, "partner_id": _SHP_PARTNER_ID}).encode()
    req = _urllib_req.Request(url, data=payload, method="POST", headers={"Content-Type": "application/json"})
    try:
        with _urllib_req.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
            tokens = _shp_load_tokens()
            if str(shop_id) in tokens:
                tokens[str(shop_id)].update(data)
                _shp_save_tokens(tokens)
            return data, None
    except _urllib_err.HTTPError as e:
        try: body = e.read().decode()
        except: body = ""
        return None, f"HTTP {e.code}: {body}"
    except Exception as e:
        return None, str(e)


def _shp_api(shop_id, path, params=None, retry=True):
    tokens = _shp_load_tokens()
    t = tokens.get(str(shop_id), {})
    access_token = t.get("access_token")
    if not access_token:
        return None, "Loja nÃ£o conectada"
    ts = int(_time.time())
    sign = _shp_sign(path, ts, access_token, str(shop_id))
    query = {
        "partner_id": _SHP_PARTNER_ID, "timestamp": ts,
        "access_token": access_token, "shop_id": shop_id, "sign": sign,
    }
    if params:
        query.update(params)
    url = _SHP_BASE + path + "?" + _urllib_parse.urlencode(query)
    req = _urllib_req.Request(url, headers={"Content-Type": "application/json"})
    try:
        with _urllib_req.urlopen(req, timeout=15) as r:
            return json.loads(r.read()), None
    except _urllib_err.HTTPError as e:
        try: body = e.read().decode()
        except: body = ""
        if retry:
            rf = t.get("refresh_token")
            if rf:
                new_data, _ = _shp_refresh_token(shop_id, rf)
                if new_data:
                    return _shp_api(shop_id, path, params, retry=False)
        return None, f"HTTP {e.code}: {body}"
    except Exception as e:
        return None, str(e)


@app.route("/api-teste/shopee/auth")
def api_teste_shopee_auth():
    path = "/api/v2/shop/auth_partner"
    ts = int(_time.time())
    sign = _shp_sign(path, ts)
    base_str = f"{_SHP_PARTNER_ID}{path}{ts}"
    (BASE_DIR / "shp_debug.log").write_text(
        f"auth: partner_id={_SHP_PARTNER_ID} ts={ts} base={base_str} sign={sign}", encoding="utf-8")
    params = _urllib_parse.urlencode({
        "partner_id": _SHP_PARTNER_ID, "timestamp": ts, "sign": sign,
        "redirect": _SHP_REDIRECT,
    })
    return redirect(f"{_SHP_BASE}{path}?{params}")


@app.route("/api-teste/shopee/callback")
def api_teste_shopee_callback():
    code    = request.args.get("code")
    shop_id = request.args.get("shop_id")
    error   = request.args.get("error")
    if error or not code or not shop_id:
        (BASE_DIR / "shp_debug.log").write_text(
            f"callback: error={error} code={code} shop_id={shop_id}", encoding="utf-8")
        return redirect("/?shp_err=1")
    shop_id = int(shop_id)
    path = "/api/v2/auth/token/get"
    ts = int(_time.time())
    sign = _shp_sign(path, ts)
    url = _SHP_BASE + path + "?" + _urllib_parse.urlencode({
        "partner_id": _SHP_PARTNER_ID, "timestamp": ts, "sign": sign,
    })
    payload = json.dumps({"code": code, "shop_id": shop_id, "partner_id": _SHP_PARTNER_ID}).encode()
    req = _urllib_req.Request(url, data=payload, method="POST", headers={"Content-Type": "application/json"})
    try:
        with _urllib_req.urlopen(req, timeout=10) as r:
            token_data = json.loads(r.read())
    except _urllib_err.HTTPError as e:
        try: body = e.read().decode()
        except: body = ""
        (BASE_DIR / "shp_debug.log").write_text(f"token error: HTTP {e.code} {body}", encoding="utf-8")
        return redirect("/?shp_err=1")
    except Exception as e:
        (BASE_DIR / "shp_debug.log").write_text(f"token error: {e}", encoding="utf-8")
        return redirect("/?shp_err=1")
    tokens = _shp_load_tokens()
    token_data["shop_id"] = shop_id
    tokens[str(shop_id)] = token_data
    _shp_save_tokens(tokens)
    shop_info, _ = _shp_api(shop_id, "/api/v2/shop/get_shop_info")
    if shop_info and shop_info.get("response"):
        tokens[str(shop_id)]["shop_name"] = shop_info["response"].get("shop_name", str(shop_id))
        _shp_save_tokens(tokens)
    (BASE_DIR / "shp_debug.log").write_text(f"ok: shop_id={shop_id}", encoding="utf-8")
    return redirect("/?shp_ok=1")


@app.route("/api-teste/shopee/accounts")
def api_teste_shopee_accounts():
    tokens = _shp_load_tokens()
    accounts = [{"shop_id": sid, "shop_name": t.get("shop_name", f"Loja {sid}")}
                for sid, t in tokens.items()]
    return jsonify({"accounts": accounts})


@app.route("/api-teste/shopee/disconnect", methods=["POST"])
def api_teste_shopee_disconnect():
    sid = str(request.json.get("shop_id", ""))
    tokens = _shp_load_tokens()
    tokens.pop(sid, None)
    _shp_save_tokens(tokens)
    return jsonify({"ok": True})


@app.route("/api-teste/shopee/data/shop")
def api_teste_shopee_data_shop():
    sid = int(request.args.get("shop_id"))
    d, e = _shp_api(sid, "/api/v2/shop/get_shop_info")
    if e: return jsonify({"error": e}), 400
    return jsonify(d)


@app.route("/api-teste/shopee/data/orders")
def api_teste_shopee_data_orders():
    sid = int(request.args.get("shop_id"))
    ts_now  = int(_time.time())
    ts_from = ts_now - 30 * 24 * 3600
    d, e = _shp_api(sid, "/api/v2/order/get_order_list", {
        "time_range_field": "create_time",
        "time_from": ts_from, "time_to": ts_now, "page_size": 20,
    })
    if e: return jsonify({"error": e}), 400
    return jsonify(d)


@app.route("/api-teste/shopee/data/products")
def api_teste_shopee_data_products():
    sid = int(request.args.get("shop_id"))
    d, e = _shp_api(sid, "/api/v2/product/get_item_list", {
        "offset": 0, "page_size": 20, "item_status": "NORMAL",
    })
    if e: return jsonify({"error": e}), 400
    return jsonify(d)


@app.route("/api-teste/shopee/data/finance")
def api_teste_shopee_data_finance():
    sid = int(request.args.get("shop_id"))
    d, e = _shp_api(sid, "/api/v2/payment/get_wallet_balance")
    if e:
        ts_now  = int(_time.time())
        ts_from = ts_now - 30 * 24 * 3600
        d2, e2 = _shp_api(sid, "/api/v2/payment/get_payment_list", {
            "page_no": 1, "page_size": 20,
            "create_time_from": ts_from, "create_time_to": ts_now,
        })
        if e2: return jsonify({"error": e}), 400
        return jsonify(d2)
    return jsonify(d)


@app.route("/api-teste/shopee/data/performance")
def api_teste_shopee_data_performance():
    sid = int(request.args.get("shop_id"))
    d, e = _shp_api(sid, "/api/v2/shop_performance/get_shop_performance")
    if e: return jsonify({"error": e}), 400
    return jsonify(d)


# â”€â”€ Financeiro â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
import openpyxl as _openpyxl
import csv as _csv

_FIN_PRODUTOS_FILE  = BASE_DIR / "financeiro_produtos.json"
_FIN_HISTORICO_FILE = BASE_DIR / "financeiro_historico.json"

_FIN_LOJAS = {
    "repzilla_shopee": "Repzilla (Shopee)",
}

def _fin_load_produtos():
    if _FIN_PRODUTOS_FILE.exists():
        return json.loads(_FIN_PRODUTOS_FILE.read_text(encoding="utf-8"))
    return []

def _fin_save_produtos(p):
    _FIN_PRODUTOS_FILE.write_text(json.dumps(p, ensure_ascii=False, indent=2), encoding="utf-8")

def _fin_load_historico():
    if _FIN_HISTORICO_FILE.exists():
        return json.loads(_FIN_HISTORICO_FILE.read_text(encoding="utf-8"))
    return []

def _fin_save_historico(h):
    _FIN_HISTORICO_FILE.write_text(json.dumps(h, ensure_ascii=False, indent=2), encoding="utf-8")


def _fin_parse_extrato(file_storage):
    """LÃª o extrato de transaÃ§Ãµes da Shopee (.xlsx) e retorna resumo."""
    wb = _openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    receita_bruta = 0.0
    reembolsos    = 0.0
    compensacoes  = 0.0
    n_pedidos     = 0
    n_reembolsos  = 0

    header_found = False
    for row in rows:
        if not header_found:
            if row and row[0] == "Data":
                header_found = True
            continue
        tipo  = row[1] if len(row) > 1 else None
        valor = row[5] if len(row) > 5 else None
        if not tipo or not isinstance(valor, (int, float)):
            continue
        if tipo == "Renda do pedido":
            if valor >= 0:
                receita_bruta += valor
                n_pedidos += 1
            else:
                reembolsos += valor   # jÃ¡ negativo
                n_reembolsos += 1
        elif tipo == "Ajuste":
            compensacoes += valor

    receita_liquida = receita_bruta + reembolsos + compensacoes
    return {
        "receita_bruta":   round(receita_bruta, 2),
        "reembolsos":      round(reembolsos, 2),
        "compensacoes":    round(compensacoes, 2),
        "receita_liquida": round(receita_liquida, 2),
        "n_pedidos":       n_pedidos,
        "n_reembolsos":    n_reembolsos,
    }


def _fin_parse_ads_csv(file_storage):
    """LÃª o CSV de anÃºncios CPC da Shopee e retorna gasto total e por produto."""
    content = file_storage.read().decode("utf-8-sig", errors="replace")
    lines   = content.splitlines()

    # Encontra a linha de header (comeÃ§a com "#,")
    header_idx = None
    for i, line in enumerate(lines):
        if line.startswith("#,"):
            header_idx = i
            break
    if header_idx is None:
        return {"total_ads": 0, "por_produto": []}

    reader = _csv.DictReader(lines[header_idx:])
    total_ads  = 0.0
    por_produto = []
    for row in reader:
        nome     = row.get("Nome do AnÃºncio", "").strip()
        despesas = row.get("Despesas", "0").strip().replace(",", ".")
        itens    = row.get("Itens Vendidos", "0").strip()
        gmv      = row.get("GMV", "0").strip().replace(",", ".")
        if not nome:
            continue
        try:
            desp = float(despesas)
            total_ads += desp
            por_produto.append({
                "nome":    nome,
                "ads":     round(desp, 2),
                "itens":   int(float(itens)) if itens else 0,
                "gmv_ads": round(float(gmv), 2) if gmv else 0,
            })
        except (ValueError, TypeError):
            continue

    return {"total_ads": round(total_ads, 2), "por_produto": por_produto}


def _fin_parse_product_overview(file_storage):
    """LÃª o productoverview xlsx e retorna quantidade vendida por produto."""
    wb  = _openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
    ws  = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Localiza header
    header_idx = None
    headers    = []
    for i, row in enumerate(rows):
        if row and any(str(c or "").strip().lower() in ("product name", "nome do produto", "nome do anÃºncio", "item name") for c in row):
            header_idx = i
            headers = [str(c or "").strip() for c in row]
            break
    if header_idx is None:
        return []

    # Identifica colunas
    def col(names):
        for n in names:
            for h in headers:
                if n.lower() in h.lower():
                    return headers.index(h)
        return None

    idx_name = col(["product name", "nome do produto", "item name", "nome"])
    idx_qty  = col(["units sold", "itens vendidos", "quantidade vendida", "qty sold", "vendidos"])
    idx_rev  = col(["revenue", "receita", "gmv"])

    if idx_name is None:
        return []

    produtos = []
    for row in rows[header_idx + 1:]:
        if not row or not row[idx_name]:
            continue
        nome = str(row[idx_name]).strip()
        qty  = int(float(row[idx_qty])) if idx_qty is not None and isinstance(row[idx_qty], (int, float)) else 0
        rev  = round(float(row[idx_rev]), 2) if idx_rev is not None and isinstance(row[idx_rev], (int, float)) else 0.0
        if nome:
            produtos.append({"nome": nome, "qty": qty, "receita": rev})

    return produtos


@app.route("/financeiro/produtos", methods=["GET"])
def fin_get_produtos():
    return jsonify({"produtos": _fin_load_produtos()})


@app.route("/financeiro/produtos", methods=["POST"])
def fin_save_produtos():
    produtos = request.json.get("produtos", [])
    _fin_save_produtos(produtos)
    return jsonify({"ok": True})


@app.route("/financeiro/lojas")
def fin_lojas():
    return jsonify({"lojas": [{"id": k, "nome": v} for k, v in _FIN_LOJAS.items()]})


@app.route("/financeiro/processar", methods=["POST"])
def fin_processar():

    extrato  = request.files.get("extrato")
    ads_csv  = request.files.get("ads_csv")
    overview = request.files.get("overview")

    if not extrato or not ads_csv:
        return jsonify({"error": "Envie o extrato e o CSV de anÃºncios"}), 400

    try:
        ext_data = _fin_parse_extrato(extrato)
    except Exception as e:
        return jsonify({"error": f"Erro ao ler extrato: {e}"}), 400

    try:
        ads_data = _fin_parse_ads_csv(ads_csv)
    except Exception as e:
        return jsonify({"error": f"Erro ao ler CSV de anÃºncios: {e}"}), 400

    ov_data = []
    if overview:
        try:
            ov_data = _fin_parse_product_overview(overview)
        except Exception as e:
            ov_data = []

    return jsonify({
        "extrato":   ext_data,
        "ads":       ads_data,
        "overview":  ov_data,
        "produtos":  _fin_load_produtos(),
    })


@app.route("/financeiro/salvar", methods=["POST"])
def fin_salvar():
    data     = request.json
    historico = _fin_load_historico()
    entry = {
        "id":        datetime.now().strftime("%Y%m%d-%H%M%S"),
        "timestamp": datetime.now().isoformat(),
        "loja":      data.get("loja", ""),
        "periodo_de": data.get("periodo_de", ""),
        "periodo_ate": data.get("periodo_ate", ""),
        "receita_liquida": data.get("receita_liquida", 0),
        "total_ads":       data.get("total_ads", 0),
        "custo_produtos":  data.get("custo_produtos", 0),
        "custo_embalagem": data.get("custo_embalagem", 0),
        "lucro":           data.get("lucro", 0),
        "n_pedidos":       data.get("n_pedidos", 0),
        "detalhes":        data.get("detalhes", {}),
    }
    historico.append(entry)
    _fin_save_historico(historico)
    return jsonify({"ok": True, "id": entry["id"]})


@app.route("/financeiro/historico")
def fin_historico():
    return jsonify({"historico": list(reversed(_fin_load_historico()))})


@app.route("/financeiro/historico/delete", methods=["POST"])
def fin_historico_delete():
    eid      = request.json.get("id")
    historico = _fin_load_historico()
    historico = [e for e in historico if e.get("id") != eid]
    _fin_save_historico(historico)
    return jsonify({"ok": True})


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if __name__ == "__main__":
    import webbrowser, threading, time
    threading.Thread(target=lambda: (time.sleep(1), webbrowser.open("http://localhost:5000")),
                     daemon=True).start()
    app.run(debug=False, port=5000, use_reloader=False)
